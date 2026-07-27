"""
batch_generator.py — Sinh hàng loạt hợp đồng từ 1 template + N khách hàng.

Thiết kế:
  - BatchJob: dataclass mô tả 1 job (không lưu DB, tồn tại trong request)
  - BatchRowResult: kết quả từng dòng (customer + success/error + file path)
  - BatchResult: tổng kết toàn bộ job
  - run_batch(): hàm chính — nhận shared params + list customer_id
  - build_zip(): đóng gói tất cả file thành bytes ZIP
  - import_excel_mapping(): đọc file Excel có cột per-customer override

Lý do KHÔNG lưu BatchJob vào DB:
  V1 dùng request-response đơn giản — user submit form, server xử lý,
  trả ZIP về ngay. Không cần job queue hay async. Nếu cần async sau này
  (V2+), chỉ cần thêm model BatchJob và celery/rq.

Quy tắc số hợp đồng trong batch:
  User cung cấp prefix/pattern, hệ thống tự thêm suffix theo thứ tự.
  VD: prefix="HD-DV" → 01/ACME/2026, 02/ACME/2026, ...
  Hoặc user import Excel với cột "Số hợp đồng" riêng từng dòng.
"""

import io
import os
import re
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from typing import Optional

from sqlalchemy.orm import Session

from app.models.contract import Contract, ContractStatus
from app.models.contract_event import ContractEvent
from app.models.customer import Customer
from app.models.contract_template import ContractTemplate
from app.services.docx_renderer import render_docx, build_render_context
from app.utils.amount_to_words_vi import amount_to_words
from app.utils.date_helpers import (
    parse_date_input, compute_end_date, months_between
)
from app.utils.file_naming import make_output_filename

from app.paths import OUTPUT_BATCH_DIR

OUTPUT_DIR = str(OUTPUT_BATCH_DIR)

# ─── Data types ──────────────────────────────────────────────────────────────

@dataclass
class BatchRowResult:
    customer_id: int
    customer_code: str
    customer_name: str
    contract_number: str
    success: bool
    file_path: str = ""
    file_name: str = ""
    error: str = ""


@dataclass
class BatchResult:
    job_id: str
    template_id: int
    template_name: str
    total: int = 0
    succeeded: int = 0
    failed: int = 0
    rows: list = field(default_factory=list)      # list[BatchRowResult]
    zip_path: str = ""
    started_at: datetime = field(default_factory=datetime.now)
    finished_at: Optional[datetime] = None

    @property
    def ok(self) -> bool:
        return self.succeeded > 0

    @property
    def duration_seconds(self) -> float:
        if self.finished_at:
            return (self.finished_at - self.started_at).total_seconds()
        return 0.0


@dataclass
class ImportExcelResult:
    rows: list = field(default_factory=list)   # list of dicts per row
    errors: list = field(default_factory=list)  # (row_num, message)
    total_rows: int = 0


# ─── Public API ──────────────────────────────────────────────────────────────

def run_batch(
    db: Session,
    template_id: int,
    customer_ids: list[int],
    shared_params: dict,
    per_customer_overrides: Optional[dict[int, dict]] = None,
    contract_number_prefix: str = "",
    year: Optional[int] = None,
) -> BatchResult:
    """
    Sinh hàng loạt hợp đồng.

    Args:
        db: SQLAlchemy session
        template_id: ID của ContractTemplate
        customer_ids: Danh sách customer ID cần sinh
        shared_params: Dict các trường chung (sign_date, start_date,
                       month_count, unit_price_vat, total_amount, ...)
        per_customer_overrides: Tuỳ chọn — dict {customer_id: {field: value}}
                                 cho phép override từng KH (VD từ Excel)
        contract_number_prefix: Prefix số HĐ, VD "HD-DV"
        year: Năm cho số HĐ, mặc định năm hiện tại

    Returns:
        BatchResult với danh sách kết quả từng KH và đường dẫn file ZIP.
    """
    import uuid
    job_id = uuid.uuid4().hex[:12]
    if year is None:
        year = datetime.now().year

    # Load template
    template = db.query(ContractTemplate).filter(
        ContractTemplate.id == template_id
    ).first()
    if not template:
        raise LookupError(f"Không tìm thấy template id={template_id}")
    if not template.file_path or not os.path.exists(template.file_path):
        raise FileNotFoundError(
            f"File template không tồn tại trên disk: {template.file_path!r}"
        )

    result = BatchResult(
        job_id=job_id,
        template_id=template_id,
        template_name=template.name,
        total=len(customer_ids),
    )

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    overrides = per_customer_overrides or {}

    for idx, cust_id in enumerate(customer_ids, start=1):
        customer = db.query(Customer).filter(Customer.id == cust_id).first()
        if not customer:
            result.rows.append(BatchRowResult(
                customer_id=cust_id, customer_code="?", customer_name="?",
                contract_number="", success=False,
                error=f"Không tìm thấy khách hàng id={cust_id}",
            ))
            result.failed += 1
            continue

        # Merge shared_params + per-customer override
        row_params = dict(shared_params)
        if cust_id in overrides:
            row_params.update(overrides[cust_id])

        # Auto-sinh pricing_description nếu chưa có (tương đương JS recalcPrice ở form lẻ)
        if not (row_params.get("pricing_description") or "").strip():
            row_params["pricing_description"] = _auto_pricing_description(customer, row_params)

        # Xác định số hợp đồng
        contract_number = (row_params.get("contract_number") or "").strip()
        if not contract_number:
            # Sinh số HĐ đúng quy chuẩn: bộ đếm riêng per (customer, khuôn, year)
            from app.services.contract_service import (
                get_next_contract_seq, make_contract_slug,
                build_contract_number, resolve_number_format,
            )
            number_format = resolve_number_format(db, row_params.get("product_id"))
            seq  = get_next_contract_seq(db, customer.id,
                                         number_format=number_format, year=year)
            slug = make_contract_slug(customer.legal_name or "")
            contract_number = build_contract_number(seq, slug, number_format, year)

        row = BatchRowResult(
            customer_id=cust_id,
            customer_code=customer.code,
            customer_name=customer.short_name or customer.legal_name,
            contract_number=contract_number,
            success=False,
        )

        try:
            # Kiểm tra trùng số hợp đồng
            existing = db.query(Contract).filter(
                Contract.contract_number == contract_number
            ).first()
            if existing:
                raise ValueError(f"Số HĐ '{contract_number}' đã tồn tại")

            # Build contract object (không persist vào DB ngay)
            contract = _build_contract_from_params(
                row_params, customer, template, contract_number
            )
            db.add(contract)
            db.flush()   # flush để có ID

            # Build render context
            ctx = build_render_context(contract, customer, template=template, db=db)
            contract.render_data = ctx

            # Render docx
            rendered = render_docx(template.file_path, ctx)

            # Save file
            fname = make_output_filename(
                contract_number=contract_number,
                customer_code=customer.code,
                contract_type=template.contract_type,
            )
            fpath = os.path.join(OUTPUT_DIR, fname)
            with open(fpath, "wb") as f:
                f.write(rendered)

            # Update contract record
            contract.output_file_path = fpath
            contract.output_file_name = fname
            contract.status = ContractStatus.Generated
            contract.is_batch = True
            contract.batch_job_id = job_id

            # Event
            ev = ContractEvent(
                contract_id=contract.id,
                event_type="generated",
                description=f"[Batch {job_id}] Sinh file {fname}",
                actor="batch",
            )
            db.add(ev)
            db.commit()

            row.success = True
            row.file_path = fpath
            row.file_name = fname
            result.succeeded += 1

        except Exception as e:
            db.rollback()
            # Xóa file đã ghi nếu commit fail để tránh file rác trên disk
            if fpath and os.path.exists(fpath):
                try:
                    os.remove(fpath)
                except OSError:
                    pass
            row.error = str(e)
            result.failed += 1

        result.rows.append(row)

    # Build ZIP chứa tất cả file thành công
    if result.succeeded > 0:
        zip_fname = f"batch_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        zip_path = os.path.join(OUTPUT_DIR, zip_fname)
        _build_zip(
            [r.file_path for r in result.rows if r.success and r.file_path],
            zip_path
        )
        result.zip_path = zip_path

    result.finished_at = datetime.now()
    return result


def build_zip_bytes(file_paths: list[str]) -> bytes:
    """Đóng gói list file thành bytes ZIP (trả về trực tiếp, không ghi disk)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in file_paths:
            if path and os.path.exists(path):
                zf.write(path, arcname=os.path.basename(path))
    return buf.getvalue()


def import_excel_mapping(
    file_bytes: bytes,
    customer_lookup: dict[str, int],   # code hoặc tên ngắn → customer_id
) -> ImportExcelResult:
    """
    Đọc file Excel có header row, trích xuất per-customer override.

    Cột bắt buộc: "Mã khách hàng" HOẶC "Tên khách hàng"
    Cột tùy chọn: bất kỳ field nào trong BATCH_EXCEL_COLUMNS

    Returns: ImportExcelResult với list dicts
             mỗi dict: {"customer_id": int, "contract_number": str, ...}
    """
    result = ImportExcelResult()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        result.errors.append((0, f"Không đọc được file Excel: {e}"))
        return result

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        result.errors.append((0, "File Excel trống hoặc thiếu dữ liệu."))
        return result

    # Detect header
    header_row = None
    header_idx = None
    for i, row in enumerate(rows[:5]):
        normalized = {
            str(c).strip().lower() if c else "": j
            for j, c in enumerate(row)
        }
        if "mã khách hàng" in normalized or "tên khách hàng" in normalized:
            header_row = normalized
            header_idx = i
            break

    if header_row is None:
        result.errors.append((0, "Không tìm thấy header. Cần cột 'Mã khách hàng' hoặc 'Tên khách hàng'."))
        return result

    # Map header names to field names
    col_map: dict[str, int] = {}
    for col_label, field_name in BATCH_EXCEL_COLUMNS.items():
        if col_label in header_row:
            col_map[field_name] = header_row[col_label]

    data_rows = rows[header_idx + 1:]
    result.total_rows = 0

    for row_offset, row in enumerate(data_rows):
        row_num = header_idx + 2 + row_offset

        # Skip fully empty
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        result.total_rows += 1

        # Extract row data
        row_data: dict[str, str] = {}
        for field_name, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                s = str(val).strip()
                if s:
                    row_data[field_name] = s

        # Resolve customer_id
        cust_code = row_data.pop("customer_code", "").strip()
        cust_name = row_data.pop("customer_name", "").strip()

        cust_id = None
        for key in [cust_code, cust_name]:
            if key and key in customer_lookup:
                cust_id = customer_lookup[key]
                break

        if cust_id is None:
            result.errors.append((
                row_num,
                f"Không tìm thấy KH với mã/tên: {cust_code!r} / {cust_name!r}"
            ))
            continue

        row_data["customer_id"] = cust_id
        result.rows.append(row_data)

    return result


def make_batch_template_xlsx() -> bytes:
    """Tạo file Excel mẫu cho batch import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Import"

    headers = [
        "Mã khách hàng", "Số hợp đồng", "Ngày ký",
        "Ngày bắt đầu", "Số tháng", "Số đơn vị",
        "Đơn giá có VAT", "Tổng giá trị", "Mô tả giá",
        "Đại diện Bên B", "Số uỷ quyền", "Ngày uỷ quyền",
    ]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="1A2238")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sample rows
    ws.append(["KH-001", "01/ACME/2026", "15/01/2025",
               "15/01/2025", "12", "1",
               "24000000", "24000000", "",
               "Nguyễn Văn A", "", ""])
    ws.append(["KH-002", "02/ACME/2026", "15/01/2025",
               "15/01/2025", "12", "1",
               "24000000", "24000000", "",
               "Nguyễn Văn A", "", ""])

    widths = [14, 22, 12, 12, 10, 10, 15, 15, 20, 20, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Constants ───────────────────────────────────────────────────────────────

BATCH_EXCEL_COLUMNS: dict[str, str] = {
    "mã khách hàng":       "customer_code",
    "tên khách hàng":      "customer_name",
    "số hợp đồng":         "contract_number",
    "ngày ký":             "sign_date",
    "ngày bắt đầu":        "start_date",
    "số tháng":            "month_count",
    "số đơn vị":           "unit_count",
    "đơn giá có vat":      "unit_price_vat",
    "tổng giá trị":        "total_amount",
    "mô tả giá":           "pricing_description",
    "đại diện bên b":      "party_b_representative",
    "số uỷ quyền":         "party_b_authorization_number",
    "ngày uỷ quyền":       "party_b_authorization_date",
}


# ─── Internal helpers ────────────────────────────────────────────────────────

def _auto_pricing_description(customer: Customer, params: dict, names=None) -> str:
    """
    Tự sinh mô tả cách tính giá — tương đương với JS recalcPrice() ở form.html.
    Format: "Số tiền = Đơn giá có VAT (X VNĐ) × N cơ sở (TYT A; Điểm B; ...) × M tháng = TOTAL VNĐ"

    names: danh sách cơ sở được chọn cho HĐ (override). Nếu None → dùng toàn bộ
           cơ sở tính phí của khách hàng (customer.unit_names_list()).
    """
    from app.services.contract_service import _safe_int, _safe_decimal

    unit_price_vat = _safe_decimal(params.get("unit_price_vat"))
    month_count    = _safe_int(params.get("month_count"))

    # Danh sách tên: cơ sở đã chọn (nếu có) hoặc toàn bộ cơ sở tính phí
    names = list(names) if names else customer.unit_names_list()
    unit_count = _safe_int(params.get("unit_count")) or (len(names) if names else customer.total_units)

    if not unit_price_vat or not month_count:
        return ""

    total = unit_price_vat * Decimal(unit_count) * Decimal(month_count)

    unit_detail = f" ({'; '.join(names)})" if len(names) > 1 else ""

    def _fmt(n) -> str:
        return f"{int(n):,}".replace(",", ".")

    return (
        f"Số tiền = Đơn giá có VAT ({_fmt(unit_price_vat)} VNĐ) "
        f"× {unit_count} cơ sở{unit_detail} "
        f"× {month_count} tháng = {_fmt(total)} VNĐ"
    )


def _build_contract_from_params(
    params: dict,
    customer: Customer,
    template: ContractTemplate,
    contract_number: str,
) -> Contract:
    """Xây dựng Contract object từ params dict. Chưa add vào session."""
    from app.services.contract_service import (
        _safe_parse_date, _safe_int, _safe_decimal
    )

    sign_date  = _safe_parse_date(params.get("sign_date"))
    start_date = _safe_parse_date(params.get("start_date"))
    end_date   = _safe_parse_date(params.get("end_date"))
    month_count = _safe_int(params.get("month_count"))

    if month_count and start_date and not end_date:
        end_date = compute_end_date(start_date, month_count)
    if month_count is None and start_date and end_date:
        month_count = months_between(start_date, end_date)

    unit_count = _safe_int(params.get("unit_count")) or 1
    unit_price_pre_vat = _safe_decimal(params.get("unit_price_pre_vat"))
    vat_rate = _safe_decimal(params.get("vat_rate")) or Decimal("0")
    unit_price_vat = _safe_decimal(params.get("unit_price_vat"))
    total_amount = _safe_decimal(params.get("total_amount"))

    if unit_price_pre_vat and unit_price_vat is None:
        unit_price_vat = unit_price_pre_vat * (1 + vat_rate / 100)

    quantity = _safe_decimal(params.get("quantity"))
    if quantity is None and month_count:
        quantity = Decimal(unit_count) * Decimal(month_count)

    if total_amount is None and unit_price_vat and quantity:
        total_amount = unit_price_vat * quantity

    amount_text = (params.get("amount_text") or "").strip()
    if not amount_text and total_amount:
        try:
            amount_text = amount_to_words(total_amount)
        except Exception:
            amount_text = ""

    return Contract(
        contract_number=contract_number,
        status=ContractStatus.Draft,
        customer_id=customer.id,
        template_id=template.id,
        product_id=_safe_int(params.get("product_id")),
        sign_date=sign_date,
        start_date=start_date,
        end_date=end_date,
        month_count=month_count,
        sign_location=(params.get("sign_location") or "").strip() or None,
        unit_count=unit_count,
        quantity=quantity,
        unit_price_pre_vat=unit_price_pre_vat,
        vat_rate=vat_rate,
        unit_price_vat=unit_price_vat,
        total_amount=total_amount,
        amount_text=amount_text or None,
        pricing_description=(params.get("pricing_description") or "").strip() or None,
        party_b_representative=(params.get("party_b_representative") or "").strip() or None,
        party_b_authorization_number=(params.get("party_b_authorization_number") or "").strip() or None,
        party_b_authorization_date=(params.get("party_b_authorization_date") or "").strip() or None,
        notes=(params.get("notes") or "").strip() or None,
    )


def _build_zip(file_paths: list[str], zip_path: str) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in file_paths:
            if path and os.path.exists(path):
                zf.write(path, arcname=os.path.basename(path))

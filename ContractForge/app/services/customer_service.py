"""
CustomerService — toàn bộ business logic cho module Khách hàng.

Chức năng:
  - list_customers: lọc + phân trang
  - get_customer / get_by_code: tra cứu đơn lẻ
  - create_customer / update_customer: tạo & cập nhật
  - toggle_active: bật/tắt trạng thái
  - get_filter_options: lấy danh sách ward, type để render dropdown
  - import_from_excel: đọc file .xlsx, trả về kết quả chi tiết
  - generate_next_code: tự sinh mã KH-NNN tiếp theo
"""

import io
import re
from dataclasses import dataclass, field
from typing import Optional

from sqlalchemy.orm import Session
from sqlalchemy import or_, func

from app.models.customer import Customer
from app.models.customer_unit import CustomerUnit, UNIT_TYPES


# ---------------------------------------------------------------------------
# Data-transfer objects
# ---------------------------------------------------------------------------

@dataclass
class CustomerFilters:
    q: str = ""                     # full-text search
    ward: str = ""                  # lọc theo xã / phường
    customer_type: str = ""         # lọc theo loại đơn vị
    active_only: bool = True        # mặc định chỉ hiện khách đang hoạt động
    page: int = 1
    per_page: int = 25
    sort_by: str = "ward"           # cột sắp xếp mặc định
    sort_dir: str = "asc"           # "asc" | "desc"


@dataclass
class ImportResult:
    total_rows: int = 0
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list = field(default_factory=list)   # list of (row_num, message)
    warnings: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Core service functions
# ---------------------------------------------------------------------------

def generate_next_code(db: Session) -> str:
    """
    Sinh mã KH-NNN tiếp theo.
    Lấy toàn bộ mã KH-* rồi lọc phía Python — tránh lỗi khi có mã
    không phải số như KH-TEST-EXPLICIT khiến ORDER BY string trả sai.
    """
    all_codes = (
        db.query(Customer.code)
        .filter(Customer.code.like("KH-%"))
        .all()
    )
    max_num = 0
    for (code,) in all_codes:
        m = re.match(r"^KH-(\d+)$", code)
        if m:
            max_num = max(max_num, int(m.group(1)))
    return f"KH-{max_num + 1:03d}"


def list_customers(db: Session, filters: CustomerFilters):
    """
    Trả về (rows, total_count) sau khi lọc + phân trang.
    Mỗi phần tử của rows là (Customer, contract_count: int) để tránh
    lazy-load sau khi session đóng.
    """
    from datetime import date as _date, timedelta as _timedelta
    from sqlalchemy import func, case, and_
    from app.models.contract import Contract
    from app.services import lifecycle

    # Dùng chung định nghĩa với dashboard. Bản cũ tự liệt kê tay và thiếu
    # 'Invoiced', nên cùng một khách hàng ra số hợp đồng hiệu lực khác nhau
    # tuỳ xem bạn đang nhìn dashboard hay danh sách khách hàng.
    _ACTIVE_ST = lifecycle.ACTIVE_LIFECYCLE_STATUS_VALUES
    _today = _date.today()
    _soon = _today + _timedelta(days=lifecycle.EXPIRING_SOON_DAYS)
    q = (
        db.query(
            Customer,
            func.count(Contract.id).label("contract_count"),
            func.sum(case((Contract.status.in_(_ACTIVE_ST), 1), else_=0)).label("active_count"),
            # "Sắp hết hạn" tính từ end_date, không còn là một trạng thái.
            func.sum(case((and_(
                Contract.status.in_(_ACTIVE_ST),
                Contract.end_date.isnot(None),
                Contract.end_date > _today,
                Contract.end_date <= _soon,
            ), 1), else_=0)).label("expiring_count"),
        )
        .outerjoin(Contract, Contract.customer_id == Customer.id)
        .group_by(Customer.id)
    )

    q = q.filter(Customer.is_deleted == False)  # luôn loại KH đã xóa
    if filters.active_only:
        q = q.filter(Customer.is_active == True)

    if filters.q:
        kw = f"%{filters.q.strip()}%"
        q = q.filter(
            or_(
                Customer.legal_name.ilike(kw),
                Customer.short_name.ilike(kw),
                Customer.code.ilike(kw),
                Customer.representative_name.ilike(kw),
                Customer.ward.ilike(kw),
            )
        )

    if filters.ward:
        q = q.filter(Customer.ward == filters.ward)

    if filters.customer_type:
        q = q.filter(Customer.customer_type == filters.customer_type)

    # total_count: count distinct customers matching filters
    total = q.count()

    # ── Sắp xếp ──────────────────────────────────────────────────────────────
    from sqlalchemy import asc, desc, case, cast, Integer as SAInt
    _dir = desc if filters.sort_dir == "desc" else asc

    SORT_MAP = {
        "code":          Customer.code,
        "legal_name":    Customer.legal_name,
        "customer_type": Customer.customer_type,
        "ward":          Customer.ward,
        "province":      Customer.province,
        "representative_name": Customer.representative_name,
        "is_active":     Customer.is_active,
        "contract_count": func.count(Contract.id),
    }
    sort_col = SORT_MAP.get(filters.sort_by, Customer.ward)

    # Secondary sort: luôn thêm legal_name để kết quả ổn định
    if filters.sort_by == "legal_name":
        order_clause = _dir(sort_col)
    else:
        order_clause = (_dir(sort_col), asc(Customer.legal_name))

    rows = (
        q.order_by(*([order_clause] if not isinstance(order_clause, tuple) else order_clause))
        .offset((filters.page - 1) * filters.per_page)
        .limit(filters.per_page)
        .all()
    )
    return rows, total


def get_customer(db: Session, customer_id: int) -> Optional[Customer]:
    from sqlalchemy.orm import joinedload
    return (db.query(Customer)
            .options(joinedload(Customer.sub_units))
            .filter(Customer.id == customer_id)
            .first())


def get_by_code(db: Session, code: str) -> Optional[Customer]:
    return db.query(Customer).filter(Customer.code == code).first()


def create_customer(db: Session, data: dict) -> Customer:
    """
    Tạo khách hàng mới từ dict.
    Raises ValueError nếu code đã tồn tại.
    """
    code = (data.get("code") or "").strip()
    if not code:
        code = generate_next_code(db)
        data["code"] = code

    if get_by_code(db, code):
        raise ValueError(f"Mã khách hàng '{code}' đã tồn tại.")

    customer = Customer(**_clean(data))
    db.add(customer)
    db.commit()
    db.refresh(customer)
    return customer


def update_customer(db: Session, customer_id: int, data: dict) -> Customer:
    """
    Cập nhật khách hàng.
    Chỉ cập nhật các trường có trong data — không ghi đè code trừ khi được chỉ định rõ.
    Raises ValueError / LookupError khi cần.
    """
    customer = get_customer(db, customer_id)
    if not customer:
        raise LookupError(f"Không tìm thấy khách hàng id={customer_id}")

    cleaned = _clean(data)

    # Kiểm tra conflict code nếu có thay đổi code
    new_code = cleaned.get("code")
    if new_code and new_code != customer.code:
        existing = get_by_code(db, new_code)
        if existing and existing.id != customer_id:
            raise ValueError(f"Mã '{new_code}' đã được dùng bởi khách hàng khác.")

    # Không cho phép set code thành None/rỗng
    if "code" in cleaned and not cleaned["code"]:
        del cleaned["code"]

    for k, v in cleaned.items():
        setattr(customer, k, v)

    db.commit()
    db.refresh(customer)
    return customer


def toggle_active(db: Session, customer_id: int) -> Customer:
    customer = get_customer(db, customer_id)
    if not customer:
        raise LookupError(f"Không tìm thấy khách hàng id={customer_id}")
    customer.is_active = not customer.is_active
    db.commit()
    db.refresh(customer)
    return customer


def get_filter_options(db: Session) -> dict:
    """Lấy danh sách ward + customer_type duy nhất để render dropdown."""
    wards = (
        db.query(Customer.ward)
        .filter(Customer.ward.isnot(None), Customer.ward != "", Customer.is_deleted == False)
        .distinct()
        .order_by(Customer.ward)
        .all()
    )
    types = (
        db.query(Customer.customer_type)
        .filter(Customer.customer_type.isnot(None), Customer.is_deleted == False)
        .distinct()
        .order_by(Customer.customer_type)
        .all()
    )
    return {
        "wards": [r[0] for r in wards],
        "customer_types": [r[0] for r in types],
    }


# ---------------------------------------------------------------------------
# Excel Import
# ---------------------------------------------------------------------------

# Mapping: tên cột trong Excel → tên trường model
EXCEL_COLUMN_MAP = {
    "mã khách hàng":        "code",
    "tên pháp lý":          "legal_name",
    "tên ngắn":             "short_name",
    "loại đơn vị":          "customer_type",
    "địa chỉ":              "address",
    "xã / phường":          "ward",
    "xã / phường / đặc khu": "ward",
    "tỉnh / thành phố":     "province",
    "tỉnh / tp":            "province",
    "huyện / quận":         "district",
    "mã số thuế":           "tax_code",
    "đại diện pháp lý":     "representative_name",
    "chức vụ":              "representative_title",
    "số tài khoản":         "bank_account",
    "ngân hàng":            "bank_name",
    "tên thụ hưởng":        "beneficiary_name",
    "đầu mối liên hệ":      "contact_name",
    "điện thoại":           "phone",
    "email":                "email",
    "ghi chú":              "notes",
    "mã cơ sở kcb":        "kcb_code",
    "mã kcb bhyt":          "kcb_code",
    "số quyết định thành lập": "establishment_decision",
    "số qđ thành lập":      "establishment_decision",
    "ngày thành lập":       "establishment_date",
    "giới tính đại diện":   "representative_gender",
    "giới tính":            "representative_gender",
}

def import_from_excel(db: Session, file_bytes: bytes, filename: str) -> ImportResult:
    """
    Đọc file Excel và upsert vào DB.
    - Nếu 'mã khách hàng' trống → tự sinh mã mới (insert).
    - Nếu 'mã khách hàng' đã có → update bản ghi hiện tại.
    - Bỏ qua row hoàn toàn trống.
    - Ghi nhận lỗi nhưng không dừng toàn bộ import.
    """
    result = ImportResult()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        result.errors.append((0, f"Không đọc được file Excel: {e}"))
        return result

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result.errors.append((0, "File Excel trống."))
        return result

    # --- Detect header row ---
    header_row_idx = None
    col_map = {}
    for i, row in enumerate(rows[:5]):          # tìm header trong 5 dòng đầu
        normalized = {str(c).strip().lower() if c else "": j for j, c in enumerate(row)}
        hits = {k: v for k, v in normalized.items() if k in EXCEL_COLUMN_MAP}
        if hits:
            header_row_idx = i
            col_map = {EXCEL_COLUMN_MAP[k]: v for k, v in hits.items()}
            break

    if header_row_idx is None:
        result.errors.append((0, (
            "Không tìm thấy hàng tiêu đề hợp lệ trong 5 dòng đầu. "
            "Vui lòng dùng file mẫu Excel với đúng tên cột."
        )))
        return result

    # Kiểm tra cột bắt buộc (theo field name đã map)
    if "legal_name" not in col_map:
        result.errors.append((0, "Thiếu cột bắt buộc 'Tên pháp lý' trong file Excel."))
        return result

    data_rows = rows[header_row_idx + 1:]
    result.total_rows = len(data_rows)

    for row_offset, row in enumerate(data_rows):
        row_num = header_row_idx + 2 + row_offset  # 1-based cho người dùng

        # Bỏ qua row hoàn toàn trống
        if all(c is None or str(c).strip() == "" for c in row):
            result.total_rows -= 1
            continue

        # Trích xuất data từ row theo col_map
        data = {}
        for field_name, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                val = str(val).strip()
                if val:
                    data[field_name] = val

        # Validate: tên pháp lý không được trống
        if not data.get("legal_name"):
            result.errors.append((row_num, "Thiếu 'Tên pháp lý' — bỏ qua dòng này."))
            result.skipped += 1
            continue

        # Mặc định province = Đồng Nai nếu không có
        data.setdefault("province", "Đồng Nai")
        # Chuẩn hóa dữ liệu theo mô hình hành chính 2 cấp: không dùng cấp huyện mới
        data.pop("district", None)
        data.setdefault("customer_type", "Trạm y tế")
        data.setdefault("is_active", True)

        code = data.get("code", "").strip()

        try:
            if code:
                existing = get_by_code(db, code)
                if existing:
                    for k, v in _clean(data).items():
                        setattr(existing, k, v)
                    db.commit()
                    result.updated += 1
                else:
                    customer = Customer(**_clean(data))
                    db.add(customer)
                    db.commit()
                    result.created += 1
            else:
                data["code"] = generate_next_code(db)
                customer = Customer(**_clean(data))
                db.add(customer)
                db.commit()
                result.created += 1
        except Exception as e:
            db.rollback()
            result.errors.append((row_num, f"Lỗi khi lưu: {e}"))
            result.skipped += 1

    return result


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

# Tập trường hợp lệ của model Customer
_CUSTOMER_FIELDS = {
    "code", "legal_name", "short_name", "customer_type",
    "address", "ward", "district", "province",
    "tax_code", "representative_name", "representative_title",
    "bank_account", "bank_name", "beneficiary_name",
    "contact_name", "phone", "email",
    "notes", "is_active", "budget_relation_code", "kcb_code",
    "establishment_decision", "establishment_date",
    "representative_gender",
}


def _clean(data: dict) -> dict:
    """Lọc chỉ giữ các key hợp lệ, strip string, convert is_active."""
    out = {}
    for k, v in data.items():
        if k not in _CUSTOMER_FIELDS:
            continue
        if isinstance(v, str):
            v = v.strip() or None
        if k == "is_active" and isinstance(v, str):
            v = v.lower() not in ("false", "0", "không", "no")
        out[k] = v
    return out


# ─── XÓA KHÁCH HÀNG ───────────────────────────────────────────────────────────

def _soft_delete(c: Customer) -> None:
    """Helper nội bộ: đánh dấu soft-delete."""
    from datetime import datetime
    c.is_deleted = True
    c.deleted_at = datetime.utcnow()


def delete_customer(db: Session, customer_id: int):
    """Soft-delete 1 khách hàng. Trả về Customer nếu thành công, None nếu không tìm thấy."""
    c = db.query(Customer).filter(
        Customer.id == customer_id, Customer.is_deleted == False).first()
    if not c:
        return None
    _soft_delete(c)
    db.commit()
    return c


def bulk_delete_customers(db: Session, ids: list[int]) -> int:
    """Soft-delete nhiều khách hàng. Trả về số bản ghi đã xử lý."""
    count = 0
    for cid in ids:
        c = db.query(Customer).filter(
            Customer.id == cid, Customer.is_deleted == False).first()
        if c:
            _soft_delete(c)
            count += 1
    db.commit()
    return count


def restore_customer(db: Session, customer_id: int) -> Customer:
    """Khôi phục khách hàng từ thùng rác."""
    c = db.query(Customer).filter(
        Customer.id == customer_id, Customer.is_deleted == True).first()
    if not c:
        raise LookupError(f"Không tìm thấy khách hàng đã xóa id={customer_id}")
    c.is_deleted = False
    c.deleted_at = None
    db.commit()
    db.refresh(c)
    return c


def list_deleted_customers(db: Session) -> list:
    """Danh sách khách hàng đã soft-delete, mới nhất trước."""
    return (
        db.query(Customer)
        .filter(Customer.is_deleted == True)
        .order_by(Customer.deleted_at.desc())
        .all()
    )


# ─── ĐƠN VỊ CON ────────────────────────────────────────────────────────────────

def get_sub_units(db: Session, customer_id: int) -> list:
    return (db.query(CustomerUnit)
            .filter(CustomerUnit.customer_id == customer_id)
            .order_by(CustomerUnit.sort_order, CustomerUnit.id)
            .all())


def save_sub_units(db: Session, customer_id: int, units_data: list[dict]) -> None:
    """
    Đồng bộ danh sách đơn vị con từ form data.
    units_data: list of {id?, name, unit_type, address, notes}
    """
    # Delete removed units (not in submitted data)
    submitted_ids = {int(u["id"]) for u in units_data if u.get("id")}
    existing = db.query(CustomerUnit).filter(
        CustomerUnit.customer_id == customer_id
    ).all()
    for u in existing:
        if u.id not in submitted_ids:
            db.delete(u)

    # Upsert
    for i, udata in enumerate(units_data):
        name = (udata.get("name") or "").strip()
        if not name:
            continue
        uid = udata.get("id")
        if uid:
            obj = db.query(CustomerUnit).filter(CustomerUnit.id == int(uid)).first()
        else:
            obj = None
        if obj is None:
            obj = CustomerUnit(customer_id=customer_id)
            db.add(obj)
        obj.name               = name
        obj.unit_type          = (udata.get("unit_type") or "Điểm trạm").strip()
        obj.address            = (udata.get("address") or "").strip() or None
        obj.notes              = (udata.get("notes") or "").strip() or None
        obj.is_active          = True
        obj.sort_order         = i
        # include_in_billing: mặc định True nếu không được truyền (sub_unit mới)
        if "include_in_billing" in udata:
            obj.include_in_billing = bool(udata["include_in_billing"])

    db.commit()


# ---------------------------------------------------------------------------
# Import định dạng nhóm: Trạm chính + Điểm trạm  (ds_55_trạm format)
# ---------------------------------------------------------------------------

@dataclass
class GroupedImportResult:
    total_customers: int = 0
    created: int = 0
    updated: int = 0
    total_units: int = 0           # tổng điểm trạm đã tạo/cập nhật
    skipped: int = 0
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


def import_grouped_from_excel(
    db: Session,
    file_bytes: bytes,
    filename: str,
    sheet_name: str = None,
) -> "GroupedImportResult":
    """
    Đọc file Excel định dạng nhóm:
      - Dòng 'Trạm y tế' → tạo/cập nhật Customer
      - Dòng 'Điểm trạm' → tạo CustomerUnit gắn vào trạm chính kế trước

    Cột bắt buộc phát hiện tự động: STT | Đơn vị | Loại
    Upsert theo legal_name (không phân biệt hoa/thường).
    """
    result = GroupedImportResult()

    try:
        import openpyxl
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        sheet_used = ws.title
    except Exception as e:
        result.errors.append((0, f"Không đọc được file Excel: {e}"))
        return result

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        result.errors.append((0, "File Excel trống."))
        return result

    # ── Tìm header row ──────────────────────────────────────────────────────
    HEADER_KEYWORDS = {"stt", "đơn vị", "loại"}
    header_idx = None
    col_name = col_type = col_group = None

    for i, row in enumerate(all_rows[:8]):
        norm = [str(c).strip().lower() if c is not None else "" for c in row]
        found = {kw for kw in HEADER_KEYWORDS if any(kw in cell for cell in norm)}
        if found >= HEADER_KEYWORDS:
            header_idx = i
            for j, cell in enumerate(norm):
                if "đơn vị" in cell and col_name is None:
                    col_name = j
                if "loại" in cell and col_type is None:
                    col_type = j
            col_group = (col_type + 1) if col_type is not None else 4
            break

    if header_idx is None or col_name is None or col_type is None:
        result.errors.append((0, (
            "Không tìm thấy dòng tiêu đề hợp lệ (cần có 'STT', 'Đơn vị', 'Loại'). "
            "Vui lòng dùng đúng file danh sách trạm."
        )))
        return result

    data_rows = all_rows[header_idx + 1:]

    # ── Build groups ─────────────────────────────────────────────────────────
    groups: dict = {}         # {gid: {tram_chinh, diem_tram, row_num}}
    current_group_id = None

    tram_counter = 0   # dùng bộ đếm riêng — không đọc cột group_id vì
                       # cột đó có thể chứa đơn giá (500000) thay vì số nhóm

    for row_offset, row in enumerate(data_rows):
        row_num = header_idx + 2 + row_offset

        if all(c is None or str(c).strip() == "" for c in row):
            continue

        raw_name = row[col_name] if col_name < len(row) else None
        raw_type = row[col_type] if col_type < len(row) else None

        name = str(raw_name).strip() if raw_name is not None else ""
        loai = str(raw_type).strip() if raw_type is not None else ""

        if not name:
            continue

        loai_lower = loai.lower().strip()
        if "trạm y tế" in loai_lower or "trạm y te" in loai_lower:
            tram_counter += 1
            current_group_id = tram_counter
            groups[current_group_id] = {
                "tram_chinh": name,
                "diem_tram": [],
                "row_num": row_num,
            }
        elif "điểm trạm" in loai_lower or "diem tram" in loai_lower:
            if current_group_id is not None and current_group_id in groups:
                groups[current_group_id]["diem_tram"].append(name)
            else:
                result.warnings.append((row_num, f"Điểm trạm '{name}' không xác định được trạm chính — bỏ qua."))

    if not groups:
        result.errors.append((0, "Không tìm thấy dòng 'Trạm y tế' nào trong file."))
        return result

    result.total_customers = len(groups)

    # ── Build name index để upsert ────────────────────────────────────────
    existing_by_name: dict = {
        c.legal_name.strip().lower(): c
        for c in db.query(Customer).filter(Customer.is_deleted == False).all()
    }

    for gid, grp in sorted(groups.items()):
        tram_name = grp["tram_chinh"]
        diem_list = grp["diem_tram"]
        row_num = grp.get("row_num", 0)

        try:
            key = tram_name.lower()
            existing = existing_by_name.get(key)

            if existing:
                existing.is_active = True
                if not existing.customer_type:
                    existing.customer_type = "Trạm y tế"
                if not existing.short_name:
                    existing.short_name = _make_short_name(tram_name)
                db.commit()
                customer = existing
                result.updated += 1
            else:
                customer = Customer(
                    code=generate_next_code(db),
                    legal_name=tram_name,
                    short_name=_make_short_name(tram_name),
                    customer_type="Trạm y tế",
                    province="Đồng Nai",
                    is_active=True,
                )
                db.add(customer)
                db.commit()
                db.refresh(customer)
                existing_by_name[key] = customer
                result.created += 1

            # Đồng bộ điểm trạm
            units_data = [{"name": n, "unit_type": "Điểm trạm"} for n in diem_list]
            _sync_sub_units_preserve(db, customer.id, units_data)
            result.total_units += len(diem_list)

        except Exception as e:
            db.rollback()
            result.errors.append((row_num, f"'{tram_name}': {e}"))
            result.skipped += 1

    return result


def _make_short_name(legal_name: str) -> str:
    """'Trạm y tế phường Biên Hòa' → 'TYT phường Biên Hòa'."""
    s = legal_name.strip()
    for prefix in ("Trạm y tế ", "Trạm Y tế ", "TRẠM Y TẾ ", "Trạm Y Tế "):
        if s.lower().startswith(prefix.lower()):
            return "TYT " + s[len(prefix):]
    return s


def _sync_sub_units_preserve(db: Session, customer_id: int, units_data: list) -> None:
    """
    Đồng bộ điểm trạm bảo toàn dữ liệu:
    - Tên trùng → giữ nguyên, is_active = True
    - Trong DB nhưng không trong file → is_active = False (ẩn, không xóa cứng)
    - Mới trong file → tạo thêm
    """
    existing_units = (
        db.query(CustomerUnit)
        .filter(CustomerUnit.customer_id == customer_id)
        .all()
    )
    existing_by_name = {u.name.strip().lower(): u for u in existing_units}
    incoming_lower = {d["name"].strip().lower() for d in units_data if d.get("name")}

    # Ẩn điểm trạm không còn trong file
    for u in existing_units:
        if u.name.strip().lower() not in incoming_lower:
            u.is_active = False

    # Tạo mới / kích hoạt lại
    for i, udata in enumerate(units_data):
        name = udata.get("name", "").strip()
        if not name:
            continue
        key = name.lower()
        if key in existing_by_name:
            existing_by_name[key].is_active = True
            existing_by_name[key].sort_order = i
        else:
            obj = CustomerUnit(
                customer_id=customer_id,
                name=name,
                unit_type=udata.get("unit_type", "Điểm trạm"),
                is_active=True,
                sort_order=i,
            )
            db.add(obj)

    db.commit()

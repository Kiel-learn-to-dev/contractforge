"""
Router: /contracts

GET  /contracts                  → danh sách + lọc
GET  /contracts/new              → form tạo hợp đồng mới (wizard)
POST /contracts/new              → submit form tạo draft
GET  /contracts/{id}             → chi tiết
POST /contracts/{id}/generate    → render .docx
POST /contracts/{id}/status      → đổi trạng thái
GET  /contracts/{id}/download    → tải file .docx
POST /contracts/{id}/upload-scan → upload PDF scan hợp đồng đã ký
GET  /contracts/{id}/signed-pdf  → xem PDF scan (inline)
POST /contracts/{id}/delete-scan → xóa PDF scan
"""

import io, math, os, re
from datetime import date, datetime
from fastapi import APIRouter, Request, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from typing import Optional

from app.database import SessionLocal
from app.paths import TEMPLATES_DIR, SIGNED_SCANS_DIR, INVOICE_DOCS_DIR, PAYMENT_SLIPS_DIR
from app.models.contract import Contract, ContractStatus
from app.models.contract_event import ContractEvent
from app.models.customer import Customer
from app.models.product import Product
from app.models.contract_template import ContractTemplate
from app.services.contract_service import (
    list_contracts, sum_contracts, get_contract, create_draft,
    generate_docx, update_status, get_stats, VALID_TRANSITIONS,
    make_contract_slug, get_next_contract_seq, build_contract_number,
    delete_contract, bulk_delete, bulk_update_status,
    _build_contract_query,
)
from app.services.customer_service import CustomerFilters, list_customers

router = APIRouter(prefix="/contracts", tags=["contracts"])
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

PER_PAGE = 25

STATUS_LABELS = {
    "Draft":        ("secondary", "Nháp"),
    "Generated":    ("info",      "Đã sinh file"),
    "Sent":         ("primary",   "Đã gửi"),
    "Signed":       ("success",   "Đã ký"),
    "Active":       ("success",   "Hiệu lực"),          # legacy
    "Invoiced":     ("info",      "Đã xuất hóa đơn"),
    "PaidActive":   ("success",   "Đã thanh toán"),
    "ExpiringSoon": ("warning",   "Sắp hết hạn"),
    "Expired":      ("danger",    "Đã hết hạn"),
    "Terminated":   ("dark",      "Thanh lý"),
}


def _flash(url, msg, t="success"):
    sep = "&" if "?" in url else "?"
    return RedirectResponse(f"{url}{sep}msg={msg}&msg_type={t}", status_code=303)


def _fv(request):
    return {"flash_msg": request.query_params.get("msg", ""),
            "flash_type": request.query_params.get("msg_type", "success")}


# ─── LIST ─────────────────────────────────────────────────────────────────────

@router.get("", response_class=HTMLResponse)
def contract_list(
    request: Request,
    q: str = "", status: str = "",
    customer_id: str = "0", product_id: str = "0",
    alert_filter: str = "",
    year: int = 0,
    date_from: str = "", date_to: str = "", date_field: str = "sign_date",
    page: int = 1, per_page: int = 25,
    sort_by: str = "created_at", sort_dir: str = "desc",
):
    from datetime import date as _date, datetime as _dt
    db = SessionLocal()
    try:
        _cid = int(customer_id) if str(customer_id).strip().isdigit() else 0
        _pid = int(product_id) if str(product_id).strip().isdigit() else 0
        per_page = per_page if per_page in (10, 25, 50, 100) else 25

        # Parse date strings
        _df = _dt.strptime(date_from, "%Y-%m-%d").date() if date_from else None
        _dt2 = _dt.strptime(date_to, "%Y-%m-%d").date() if date_to else None

        _filter_kw = dict(
            q=q, status=status,
            customer_id=_cid or None,
            product_id=_pid or None,
            alert_filter=alert_filter,
            year=year,
            date_from=_df, date_to=_dt2, date_field=date_field,
        )

        rows, total = list_contracts(
            db, **_filter_kw,
            page=page, per_page=per_page,
            sort_by=sort_by, sort_dir=sort_dir,
        )
        filtered_total_amount = sum_contracts(db, **_filter_kw)

        total_pages = max(1, math.ceil(total / per_page))
        today = _date.today()
        contracts_data = []
        for c in rows:
            cust = c.customer
            flag_unsent = (
                c.status.value == "Sent" and
                c.updated_at and
                (today - c.updated_at.date()).days > 10
            )
            flag_expiring = (
                c.end_date and
                c.status.value in ("Active", "Signed", "ExpiringSoon") and
                0 <= (c.end_date - today).days <= 60
            )
            contracts_data.append({
                "contract": c,
                "customer_name": cust.short_name or cust.legal_name if cust else "—",
                "flag_unsent": flag_unsent,
                "flag_expiring": flag_expiring,
                "total_amount": float(c.total_amount) if c.total_amount else 0,
            })

        statuses = [s.value for s in ContractStatus]
        products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
        customers = (
            db.query(Customer)
            .filter(Customer.is_active == True, Customer.is_deleted == False)
            .order_by(Customer.short_name)
            .all()
        )

        # Available years for year-pills (distinct sign_date years)
        from sqlalchemy import extract
        _yr_col = extract("year", Contract.sign_date)
        year_rows = (
            db.query(_yr_col)
            .filter(Contract.sign_date.isnot(None))
            .distinct()
            .order_by(_yr_col.desc())
            .all()
        )
        available_years = [int(r[0]) for r in year_rows if r[0]]

        # Alert badge counts
        _, cnt_unsent   = list_contracts(db, alert_filter="unsent_10d",  per_page=1)
        _, cnt_expiring = list_contracts(db, alert_filter="expiring_60", per_page=1)

        # Base query-string for sort/pagination links (all filters, no page/sort)
        import urllib.parse
        _bq_parts = [
            ("q", q), ("status", status),
            ("customer_id", str(_cid)), ("product_id", str(_pid)),
            ("alert_filter", alert_filter), ("year", str(year)),
            ("date_from", date_from), ("date_to", date_to),
            ("date_field", date_field), ("per_page", str(per_page)),
        ]
        base_qs = urllib.parse.urlencode([(k, v) for k, v in _bq_parts if v not in ("", "0")])
    finally:
        db.close()

    return templates.TemplateResponse(request, "contracts/list.html", {
        "request": request, "page_title": "Hợp đồng",
        "contracts_data": contracts_data, "total": total,
        "filtered_total_amount": filtered_total_amount,
        "total_pages": total_pages, "page": page,
        "per_page": per_page,
        "q": q, "sel_status": status,
        "sel_customer_id": _cid,
        "sel_product_id": _pid,
        "sel_alert": alert_filter,
        "sel_year": year,
        "sel_date_from": date_from, "sel_date_to": date_to, "sel_date_field": date_field,
        "sort_by": sort_by, "sort_dir": sort_dir,
        "statuses": statuses, "products": products, "customers": customers,
        "available_years": available_years,
        "status_labels": STATUS_LABELS,
        "cnt_unsent": cnt_unsent,
        "cnt_expiring": cnt_expiring,
        "base_qs": base_qs,
        **_fv(request),
    })


# ─── EXPORT ───────────────────────────────────────────────────────────────────

@router.get("/export")
def export_contracts(
    q: str = "", status: str = "",
    customer_id: str = "0", product_id: str = "0",
    alert_filter: str = "",
    year: int = 0,
    date_from: str = "", date_to: str = "", date_field: str = "sign_date",
    sort_by: str = "sign_date", sort_dir: str = "desc",
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt
    from fastapi.responses import StreamingResponse
    from sqlalchemy.orm import joinedload

    db = SessionLocal()
    try:
        _cid = int(customer_id) if str(customer_id).strip().isdigit() else 0
        _pid = int(product_id) if str(product_id).strip().isdigit() else 0
        _df = _dt.strptime(date_from, "%Y-%m-%d").date() if date_from else None
        _dt2 = _dt.strptime(date_to, "%Y-%m-%d").date() if date_to else None

        query = _build_contract_query(
            db, q=q, status=status,
            customer_id=_cid or None, product_id=_pid or None,
            alert_filter=alert_filter, year=year,
            date_from=_df, date_to=_dt2, date_field=date_field,
        )
        from app.services.contract_service import _SORT_MAP
        col = _SORT_MAP.get(sort_by, Contract.sign_date)
        order = col.desc() if sort_dir == "desc" else col.asc()
        rows = (
            query.options(joinedload(Contract.customer), joinedload(Contract.product))
            .order_by(order)
            .all()
        )
    finally:
        db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách hợp đồng"

    HEADER_FILL = PatternFill("solid", fgColor="D6E4F7")
    HEADER_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")

    headers = ["STT", "Số hợp đồng", "Khách hàng", "Sản phẩm",
               "Ngày ký", "Ngày hết hạn", "Giá trị (VNĐ)", "Trạng thái", "Ghi chú"]
    col_widths = [5, 30, 35, 15, 12, 12, 18, 15, 30]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 20

    def _fmt_date(d):
        return d.strftime("%d/%m/%Y") if d else ""

    def _fmt_money(v):
        if v is None:
            return ""
        try:
            return f"{int(v):,}".replace(",", ".")
        except Exception:
            return str(v)

    STATUS_VI = {s: STATUS_LABELS.get(s, ("", s))[1] for s in STATUS_LABELS}

    for i, c in enumerate(rows, 1):
        cust = c.customer
        prod = c.product
        ws.append([
            i,
            c.contract_number or "",
            (cust.short_name or cust.legal_name) if cust else "",
            prod.name if prod else "",
            _fmt_date(c.sign_date),
            _fmt_date(c.end_date),
            _fmt_money(c.total_amount),
            STATUS_VI.get(c.status.value, c.status.value),
            c.notes or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date as _date
    fname = f"hop-dong-{_date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── NEW ──────────────────────────────────────────────────────────────────────

@router.get("/new", response_class=HTMLResponse)
def new_form(request: Request, customer_id: int = 0, template_id: int = 0):
    db = SessionLocal()
    try:
        customers_rows, _ = list_customers(db, CustomerFilters(active_only=True, per_page=500, sort_by='code', sort_dir='asc'))
        customers = [c for c, *_ in customers_rows]
        products  = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
        tpls      = db.query(ContractTemplate).filter(ContractTemplate.is_active == True).order_by(ContractTemplate.name).all()
        # preselected
        sel_customer = db.query(Customer).filter(Customer.id == customer_id).first() if customer_id else None
        sel_template = db.query(ContractTemplate).filter(ContractTemplate.id == template_id).first() if template_id else None
    finally:
        db.close()

    return templates.TemplateResponse(request, "contracts/form.html", {
        "request": request, "page_title": "Sinh hợp đồng",
        "form_title": "Tạo hợp đồng mới",
        "customers": customers, "products": products, "templates": tpls,
        "sel_customer": sel_customer, "sel_template": sel_template,
        "errors": [], "prefill": {},
        "status_labels": STATUS_LABELS,
    })


@router.post("/new")
async def new_submit(request: Request):
    form = await request.form()
    data = dict(form)

    db = SessionLocal()
    try:
        contract = create_draft(db, data)
        flash_msg = "Đã tạo bản nháp hợp đồng"
        flash_type = "success"

        if contract.template_id:
            try:
                contract = generate_docx(db, contract.id)
                flash_msg = f"Đã tạo hợp đồng và sinh file {contract.output_file_name}"
            except (LookupError, FileNotFoundError, ValueError) as gen_err:
                flash_msg = f"Đã tạo bản nháp nhưng chưa sinh được file: {gen_err}"
                flash_type = "warning"
    except ValueError as e:
        # Re-render form with error
        customers_rows, _ = list_customers(db, CustomerFilters(active_only=True, per_page=500, sort_by='code', sort_dir='asc'))
        customers = [c for c, *_ in customers_rows]
        products  = db.query(Product).filter(Product.is_active == True).all()
        tpls      = db.query(ContractTemplate).filter(ContractTemplate.is_active == True).all()
        # NOTE: không gọi db.close() ở đây — finally block bên dưới sẽ xử lý
        return templates.TemplateResponse(request, "contracts/form.html", {
            "request": request, "page_title": "Sinh hợp đồng",
            "form_title": "Tạo hợp đồng mới",
            "customers": customers, "products": products, "templates": tpls,
            "sel_customer": None, "sel_template": None,
            "errors": [str(e)], "prefill": data,
            "status_labels": STATUS_LABELS,
        }, status_code=422)
    finally:
        db.close()

    return _flash(f"/contracts/{contract.id}", flash_msg, flash_type)



@router.get("/generate")
def generate_entry():
    return RedirectResponse(url="/contracts/new", status_code=302)


@router.get("/batch")
def batch_entry():
    return RedirectResponse(url="/batch", status_code=302)



# ─── CONTRACT NUMBER SUGGESTION API ──────────────────────────────────────────

@router.get("/customer-units", response_class=JSONResponse)
def get_customer_units(customer_id: int = 0):
    """
    Trả về danh sách cơ sở (trạm chính + điểm trạm đang hoạt động) của khách hàng
    để dựng checklist chọn cơ sở khi sinh hợp đồng.

    Trả về:
      {
        "total_units": <số cơ sở mặc định tính phí>,
        "unit_names":  [<tên các cơ sở mặc định tính phí>],   # backward-compat
        "facilities":  [ {"key","name","type","checked"} ... ]
      }
    """
    if not customer_id:
        return JSONResponse({"total_units": 1, "unit_names": [], "facilities": []})
    db = SessionLocal()
    try:
        from app.services.customer_service import get_customer
        c = get_customer(db, customer_id)
        if not c:
            return JSONResponse({"total_units": 1, "unit_names": [], "facilities": []})

        facilities = []
        # Trạm chính
        if c.bill_main_station:
            facilities.append({
                "key": "main",
                "name": c.short_name or c.legal_name,
                "type": "Trạm chính",
                "checked": True,
            })
        # Điểm trạm / cơ sở con đang hoạt động
        for u in c.sub_units:
            if u.is_active:
                facilities.append({
                    "key": f"u{u.id}",
                    "name": u.name,
                    "type": u.unit_type or "Điểm trạm",
                    "checked": bool(u.include_in_billing),
                })

        checked_names = [f["name"] for f in facilities if f["checked"]]
        return JSONResponse({
            "total_units": len(checked_names) or 1,
            "unit_names": checked_names,
            "facilities": facilities,
        })
    finally:
        db.close()


@router.get("/suggest-number", response_class=JSONResponse)
def suggest_contract_number(customer_id: int = 0, contract_type: str = "YTCS"):
    """
    Gợi ý số hợp đồng dựa trên khách hàng và loại hợp đồng.
    Trả về: { "number": "01/KDGP/DNI-TYTXAPHUOCAN/2026" }
    """
    db = SessionLocal()
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            return JSONResponse({"number": "", "error": "Khách hàng không tồn tại"})
        year = date.today().year
        seq  = get_next_contract_seq(db, customer_id, contract_type=contract_type, year=year)
        slug = make_contract_slug(customer.legal_name or "")
        number = build_contract_number(seq, slug, contract_type, year)
        return JSONResponse({"number": number})
    finally:
        db.close()


# ─── DETAIL ───────────────────────────────────────────────────────────────────

# ─── BULK ACTIONS ─────────────────────────────────────────────────────────────

@router.post("/bulk-action", response_class=HTMLResponse)
async def bulk_action(request: Request):
    """Xử lý chọn nhiều: xóa hoặc cập nhật trạng thái cho nhiều hợp đồng."""
    form = await request.form()
    action    = form.get("bulk_action", "")
    ids_raw   = form.getlist("selected_ids")
    new_status = form.get("bulk_status", "")

    try:
        ids = [int(i) for i in ids_raw if i.strip().isdigit()]
    except ValueError:
        ids = []

    if not ids:
        return _flash("/contracts", "Chưa chọn hợp đồng nào", "warning")

    db = SessionLocal()
    try:
        if action == "delete":
            n = bulk_delete(db, ids)
            return _flash("/contracts", f"Đã xóa {n} hợp đồng", "success")
        elif action == "update_status" and new_status:
            n = bulk_update_status(db, ids, new_status)
            label = STATUS_LABELS.get(new_status, ("", new_status))[1]
            return _flash("/contracts", f"Đã cập nhật {n} hợp đồng → {label}", "success")
        else:
            return _flash("/contracts", "Hành động không hợp lệ", "danger")
    except ValueError as e:
        return _flash("/contracts", str(e), "danger")
    finally:
        db.close()


# ─── DELETE SINGLE ─────────────────────────────────────────────────────────────

@router.post("/{contract_id}/delete")
def delete_one(contract_id: int):
    db = SessionLocal()
    try:
        ok = delete_contract(db, contract_id)
    finally:
        db.close()
    if ok:
        return _flash("/contracts", "Đã xóa hợp đồng", "success")
    return _flash("/contracts", "Không tìm thấy hợp đồng", "danger")


@router.get("/{contract_id}", response_class=HTMLResponse)
def contract_detail(request: Request, contract_id: int):
    db = SessionLocal()
    try:
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")
        customer = db.query(Customer).filter(Customer.id == c.customer_id).first()
        template = db.query(ContractTemplate).filter(ContractTemplate.id == c.template_id).first() if c.template_id else None
        product  = db.query(Product).filter(Product.id == c.product_id).first() if c.product_id else None
        events   = list(c.events)   # load while session open
        allowed_next = [s.value for s in VALID_TRANSITIONS.get(c.status, set())]
    finally:
        db.close()

    return templates.TemplateResponse(request, "contracts/detail.html", {
        "request": request, "page_title": "Hợp đồng",
        "c": c, "customer": customer, "template": template, "product": product,
        "events": events, "allowed_next": allowed_next,
        "status_labels": STATUS_LABELS,
        **_fv(request),
    })


# ─── GENERATE ─────────────────────────────────────────────────────────────────

@router.post("/{contract_id}/generate")
def do_generate(request: Request, contract_id: int):
    db = SessionLocal()
    try:
        contract = generate_docx(db, contract_id)
        fname = contract.output_file_name
    except (LookupError, FileNotFoundError, ValueError) as e:
        return _flash(f"/contracts/{contract_id}", str(e), "danger")
    finally:
        db.close()

    return _flash(f"/contracts/{contract_id}",
                  f"Đã sinh file {fname}", "success")


# ─── BIỂU MẪU 08A ─────────────────────────────────────────────────────────────

@router.get("/{contract_id}/form-08a")
def download_form_08a(contract_id: int):
    """Sinh & tải biểu Mẫu số 08a (Bảng xác định giá trị KLCV hoàn thành) cho hợp đồng."""
    db = SessionLocal()
    try:
        from app.services.form08a_service import render_form_08a
        data, fname = render_form_08a(db, contract_id)
    except (LookupError, FileNotFoundError, ValueError) as e:
        return _flash(f"/contracts/{contract_id}", f"Không sinh được biểu 08A: {e}", "danger")
    finally:
        db.close()

    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── STATUS ───────────────────────────────────────────────────────────────────

@router.post("/{contract_id}/status")
async def change_status(request: Request, contract_id: int):
    form = await request.form()
    new_status = form.get("new_status", "")
    note       = form.get("note", "")

    db = SessionLocal()
    try:
        update_status(db, contract_id, new_status, note=note)
    except (LookupError, ValueError) as e:
        return _flash(f"/contracts/{contract_id}", str(e), "danger")
    finally:
        db.close()

    return _flash(f"/contracts/{contract_id}",
                  f"Đã cập nhật trạng thái → {new_status}", "success")


# ─── SCAN PDF — UPLOAD ────────────────────────────────────────────────────────

# Các trạng thái được phép tự động chuyển sang Signed khi upload scan.
# Chỉ bao gồm những trạng thái nằm trong VALID_TRANSITIONS → Signed.
# Draft KHÔNG có trong danh sách vì Draft chỉ chuyển được sang Generated.
_AUTO_SIGN_FROM = {ContractStatus.Generated, ContractStatus.Sent}

_PDF_MAGIC = b"%PDF"
_MAX_SCAN_BYTES = 5 * 1024 * 1024  # 5 MB


@router.post("/{contract_id}/upload-scan")
async def upload_scan(contract_id: int, scan_file: UploadFile = File(...)):
    """
    Nhận file PDF scan hợp đồng đã ký.
    Validate magic bytes (không tin content-type từ client).
    Giới hạn 5 MB.
    Nếu status đang là Generated hoặc Sent → tự động chuyển sang Signed.
    """
    # 1. Đọc header để xác thực magic bytes %PDF
    header = await scan_file.read(4)
    if header != _PDF_MAGIC:
        return _flash(f"/contracts/{contract_id}",
                      "File không hợp lệ — chỉ chấp nhận PDF (.pdf)", "danger")

    # 2. Đọc phần còn lại và kiểm tra kích thước
    body = await scan_file.read()
    file_bytes = header + body
    if len(file_bytes) > _MAX_SCAN_BYTES:
        return _flash(f"/contracts/{contract_id}",
                      f"File quá lớn — tối đa 5 MB (file hiện tại: {len(file_bytes)//1024} KB)", "danger")

    db = SessionLocal()
    try:
        from app.services.contract_service import get_contract
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")

        # 3. Xác định đường dẫn file mới — deterministic, không chứa ký tự nguy hiểm
        safe_num = re.sub(r"[^\w\-]", "_", c.contract_number or str(contract_id))
        dest_name = f"scan_{contract_id}_{safe_num}.pdf"
        dest_path = SIGNED_SCANS_DIR / dest_name
        SIGNED_SCANS_DIR.mkdir(parents=True, exist_ok=True)

        # 4. Ghi file MỚI trước — nếu fail thì file cũ vẫn còn nguyên (tránh data loss)
        try:
            dest_path.write_bytes(file_bytes)
        except OSError as e:
            return _flash(f"/contracts/{contract_id}",
                          f"Không thể lưu file lên đĩa: {e}", "danger")

        # 5. Xóa file scan cũ SAU KHI ghi mới thành công
        #    Bỏ qua nếu cùng path (re-upload cùng số HĐ) — file đã bị ghi đè rồi
        old_path = c.signed_pdf_path
        if old_path and old_path != str(dest_path) and os.path.exists(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass  # File cũ không xóa được — không chặn upload mới

        # 6. Cập nhật DB — scan fields
        c.signed_pdf_path = str(dest_path)
        c.signed_pdf_name = scan_file.filename or dest_name
        c.signed_pdf_uploaded_at = datetime.utcnow()

        # 7. Auto-transition status → Signed (chỉ khi hợp lệ)
        status_note = ""
        if c.status in _AUTO_SIGN_FROM:
            old_status = c.status.value
            c.status = ContractStatus.Signed
            db.add(ContractEvent(
                contract_id=c.id,
                event_type="status_changed",
                description=f"Tự động chuyển {old_status} → Signed sau khi upload scan PDF",
                actor="system",
            ))
            status_note = " — trạng thái tự động cập nhật → Đã ký"

        # 8. Ghi event upload
        db.add(ContractEvent(
            contract_id=c.id,
            event_type="scan_uploaded",
            description=f"Upload scan PDF: {c.signed_pdf_name}",
            actor="user",
        ))

        db.commit()
    except HTTPException:
        raise
    finally:
        db.close()

    return _flash(f"/contracts/{contract_id}",
                  f"Đã upload scan PDF thành công{status_note}", "success")


# ─── SCAN PDF — XEM (inline) ──────────────────────────────────────────────────

@router.get("/{contract_id}/signed-pdf")
def view_signed_pdf(contract_id: int):
    """Phục vụ file PDF scan inline — trình duyệt sẽ mở PDF viewer."""
    db = SessionLocal()
    try:
        from app.services.contract_service import get_contract
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")
        if not c.signed_pdf_path or not os.path.exists(c.signed_pdf_path):
            raise HTTPException(404, "Chưa có file scan hoặc file không còn tồn tại trên đĩa")
        path = c.signed_pdf_path
        name = c.signed_pdf_name or os.path.basename(path)
    finally:
        db.close()

    return FileResponse(
        path=path,
        filename=name,
        media_type="application/pdf",
        content_disposition_type="inline",  # mở trong browser, không force download
    )


# ─── SCAN PDF — XÓA ──────────────────────────────────────────────────────────

@router.post("/{contract_id}/delete-scan")
def delete_scan(contract_id: int):
    """Xóa file scan PDF — xóa file trên đĩa và clear 3 cột trong DB."""
    db = SessionLocal()
    try:
        from app.services.contract_service import get_contract
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")
        if not c.signed_pdf_path:
            return _flash(f"/contracts/{contract_id}",
                          "Hợp đồng này chưa có file scan", "warning")

        # Xóa file trên đĩa
        if os.path.exists(c.signed_pdf_path):
            try:
                os.remove(c.signed_pdf_path)
            except OSError as e:
                return _flash(f"/contracts/{contract_id}",
                              f"Không thể xóa file: {e}", "danger")

        old_name = c.signed_pdf_name or ""
        c.signed_pdf_path = None
        c.signed_pdf_name = None
        c.signed_pdf_uploaded_at = None

        db.add(ContractEvent(
            contract_id=c.id,
            event_type="scan_deleted",
            description=f"Đã xóa scan PDF: {old_name}",
            actor="user",
        ))
        db.commit()
    except HTTPException:
        raise
    finally:
        db.close()

    return _flash(f"/contracts/{contract_id}", "Đã xóa file scan PDF", "success")


# ─── HÓA ĐƠN — UPLOAD ────────────────────────────────────────────────────────

_AUTO_INVOICE_FROM = {ContractStatus.Signed, ContractStatus.Active}  # Active = legacy

_MAX_INVOICE_BYTES = 5 * 1024 * 1024  # 5 MB


@router.post("/{contract_id}/upload-invoice")
async def upload_invoice(
    contract_id: int,
    invoice_file: UploadFile = File(...),
    invoice_date: Optional[str] = Form(None),
):
    """
    Upload file hóa đơn PDF.
    Nếu status đang là Signed (hoặc Active legacy) → tự động chuyển sang Invoiced.
    Yêu cầu: chỉ PDF, tối đa 5 MB.
    """
    header = await invoice_file.read(4)
    if header != _PDF_MAGIC:
        return _flash(f"/contracts/{contract_id}",
                      "File không hợp lệ — hóa đơn chỉ chấp nhận PDF (.pdf)", "danger")

    body = await invoice_file.read()
    file_bytes = header + body
    if len(file_bytes) > _MAX_INVOICE_BYTES:
        return _flash(f"/contracts/{contract_id}",
                      f"File quá lớn — tối đa 5 MB (file hiện tại: {len(file_bytes)//1024} KB)", "danger")

    db = SessionLocal()
    try:
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")

        safe_num = re.sub(r"[^\w\-]", "_", c.contract_number or str(contract_id))
        dest_name = f"invoice_{contract_id}_{safe_num}.pdf"
        dest_path = INVOICE_DOCS_DIR / dest_name
        INVOICE_DOCS_DIR.mkdir(parents=True, exist_ok=True)

        try:
            dest_path.write_bytes(file_bytes)
        except OSError as e:
            return _flash(f"/contracts/{contract_id}",
                          f"Không thể lưu file lên đĩa: {e}", "danger")

        old_path = c.invoice_pdf_path
        if old_path and old_path != str(dest_path) and os.path.exists(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass

        c.invoice_pdf_path = str(dest_path)
        c.invoice_pdf_name = invoice_file.filename or dest_name
        c.invoice_pdf_uploaded_at = datetime.utcnow()

        # Lưu ngày xuất hóa đơn nếu có
        if invoice_date:
            try:
                from datetime import date as _date
                c.invoice_date = _date.fromisoformat(invoice_date)
            except ValueError:
                pass

        status_note = ""
        if c.status in _AUTO_INVOICE_FROM:
            old_status = c.status.value
            c.status = ContractStatus.Invoiced
            db.add(ContractEvent(
                contract_id=c.id,
                event_type="status_changed",
                description=f"Tự động chuyển {old_status} → Invoiced sau khi upload hóa đơn",
                actor="system",
            ))
            status_note = " — trạng thái tự động cập nhật → Đã xuất hóa đơn"

        db.add(ContractEvent(
            contract_id=c.id,
            event_type="invoice_uploaded",
            description=f"Upload hóa đơn PDF: {c.invoice_pdf_name}",
            actor="user",
        ))
        db.commit()
    except HTTPException:
        raise
    finally:
        db.close()

    return _flash(f"/contracts/{contract_id}",
                  f"Đã upload hóa đơn thành công{status_note}", "success")


@router.get("/{contract_id}/invoice-pdf")
def view_invoice_pdf(contract_id: int):
    """Phục vụ file hóa đơn PDF inline."""
    db = SessionLocal()
    try:
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")
        if not c.invoice_pdf_path or not os.path.exists(c.invoice_pdf_path):
            raise HTTPException(404, "Chưa có file hóa đơn hoặc file không còn tồn tại")
        path = c.invoice_pdf_path
        name = c.invoice_pdf_name or os.path.basename(path)
    finally:
        db.close()

    return FileResponse(path=path, filename=name,
                        media_type="application/pdf",
                        content_disposition_type="inline")


@router.post("/{contract_id}/delete-invoice")
def delete_invoice(contract_id: int):
    """Xóa file hóa đơn PDF."""
    db = SessionLocal()
    try:
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")
        if not c.invoice_pdf_path:
            return _flash(f"/contracts/{contract_id}",
                          "Hợp đồng này chưa có file hóa đơn", "warning")

        if os.path.exists(c.invoice_pdf_path):
            try:
                os.remove(c.invoice_pdf_path)
            except OSError as e:
                return _flash(f"/contracts/{contract_id}",
                              f"Không thể xóa file: {e}", "danger")

        old_name = c.invoice_pdf_name or ""
        c.invoice_pdf_path = None
        c.invoice_pdf_name = None
        c.invoice_pdf_uploaded_at = None

        db.add(ContractEvent(
            contract_id=c.id,
            event_type="invoice_deleted",
            description=f"Đã xóa hóa đơn PDF: {old_name}",
            actor="user",
        ))
        db.commit()
    except HTTPException:
        raise
    finally:
        db.close()

    return _flash(f"/contracts/{contract_id}", "Đã xóa file hóa đơn", "success")


# ─── ỦY NHIỆM CHI — UPLOAD ────────────────────────────────────────────────────

_AUTO_PAID_FROM = {ContractStatus.Invoiced}

_MAX_SLIP_BYTES = 5 * 1024 * 1024  # 5 MB
_JPG_MAGIC  = b"\xff\xd8\xff"
_PNG_MAGIC  = b"\x89PNG"


def _is_valid_slip_file(header: bytes) -> tuple[bool, str]:
    """Kiểm tra magic bytes: PDF, JPG hoặc PNG. Trả về (valid, mime_type)."""
    if header[:4] == _PDF_MAGIC:
        return True, "application/pdf"
    if header[:3] == _JPG_MAGIC:
        return True, "image/jpeg"
    if header[:4] == _PNG_MAGIC:
        return True, "image/png"
    return False, ""


@router.post("/{contract_id}/upload-payment-slip")
async def upload_payment_slip(
    contract_id: int,
    slip_file: UploadFile = File(...),
    payment_date: Optional[str] = Form(None),
):
    """
    Upload file ủy nhiệm chi (PDF hoặc ảnh JPG/PNG).
    Nếu status đang là Invoiced → tự động chuyển sang PaidActive.
    Yêu cầu: PDF/JPG/PNG, tối đa 5 MB.
    """
    header = await slip_file.read(4)
    valid, mime = _is_valid_slip_file(header)
    if not valid:
        return _flash(f"/contracts/{contract_id}",
                      "File không hợp lệ — chỉ chấp nhận PDF, JPG hoặc PNG", "danger")

    body = await slip_file.read()
    file_bytes = header + body
    if len(file_bytes) > _MAX_SLIP_BYTES:
        return _flash(f"/contracts/{contract_id}",
                      f"File quá lớn — tối đa 5 MB (file hiện tại: {len(file_bytes)//1024} KB)", "danger")

    db = SessionLocal()
    try:
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")

        ext = ".pdf" if mime == "application/pdf" else (".jpg" if mime == "image/jpeg" else ".png")
        safe_num = re.sub(r"[^\w\-]", "_", c.contract_number or str(contract_id))
        dest_name = f"payment_{contract_id}_{safe_num}{ext}"
        dest_path = PAYMENT_SLIPS_DIR / dest_name
        PAYMENT_SLIPS_DIR.mkdir(parents=True, exist_ok=True)

        try:
            dest_path.write_bytes(file_bytes)
        except OSError as e:
            return _flash(f"/contracts/{contract_id}",
                          f"Không thể lưu file lên đĩa: {e}", "danger")

        old_path = c.payment_slip_path
        if old_path and old_path != str(dest_path) and os.path.exists(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass

        c.payment_slip_path = str(dest_path)
        c.payment_slip_name = slip_file.filename or dest_name
        c.payment_slip_uploaded_at = datetime.utcnow()

        if payment_date:
            try:
                from datetime import date as _date
                c.payment_date = _date.fromisoformat(payment_date)
            except ValueError:
                pass

        status_note = ""
        if c.status in _AUTO_PAID_FROM:
            old_status = c.status.value
            c.status = ContractStatus.PaidActive
            db.add(ContractEvent(
                contract_id=c.id,
                event_type="status_changed",
                description=f"Tự động chuyển {old_status} → PaidActive sau khi upload ủy nhiệm chi",
                actor="system",
            ))
            status_note = " — trạng thái tự động cập nhật → Đã thanh toán"

        db.add(ContractEvent(
            contract_id=c.id,
            event_type="payment_slip_uploaded",
            description=f"Upload ủy nhiệm chi: {c.payment_slip_name}",
            actor="user",
        ))
        db.commit()
    except HTTPException:
        raise
    finally:
        db.close()

    return _flash(f"/contracts/{contract_id}",
                  f"Đã upload ủy nhiệm chi thành công{status_note}", "success")


@router.get("/{contract_id}/payment-slip")
def view_payment_slip(contract_id: int):
    """Phục vụ file ủy nhiệm chi inline."""
    db = SessionLocal()
    try:
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")
        if not c.payment_slip_path or not os.path.exists(c.payment_slip_path):
            raise HTTPException(404, "Chưa có file ủy nhiệm chi hoặc file không còn tồn tại")
        path = c.payment_slip_path
        name = c.payment_slip_name or os.path.basename(path)
        ext = os.path.splitext(name)[1].lower()
        media = "application/pdf" if ext == ".pdf" else ("image/jpeg" if ext in (".jpg", ".jpeg") else "image/png")
    finally:
        db.close()

    return FileResponse(path=path, filename=name,
                        media_type=media,
                        content_disposition_type="inline")


@router.post("/{contract_id}/delete-payment-slip")
def delete_payment_slip(contract_id: int):
    """Xóa file ủy nhiệm chi."""
    db = SessionLocal()
    try:
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")
        if not c.payment_slip_path:
            return _flash(f"/contracts/{contract_id}",
                          "Hợp đồng này chưa có file ủy nhiệm chi", "warning")

        if os.path.exists(c.payment_slip_path):
            try:
                os.remove(c.payment_slip_path)
            except OSError as e:
                return _flash(f"/contracts/{contract_id}",
                              f"Không thể xóa file: {e}", "danger")

        old_name = c.payment_slip_name or ""
        c.payment_slip_path = None
        c.payment_slip_name = None
        c.payment_slip_uploaded_at = None

        db.add(ContractEvent(
            contract_id=c.id,
            event_type="payment_slip_deleted",
            description=f"Đã xóa ủy nhiệm chi: {old_name}",
            actor="user",
        ))
        db.commit()
    except HTTPException:
        raise
    finally:
        db.close()

    return _flash(f"/contracts/{contract_id}", "Đã xóa file ủy nhiệm chi", "success")


# ─── DOWNLOAD ─────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/download")
def download(contract_id: int):
    db = SessionLocal()
    try:
        c = get_contract(db, contract_id)
        if not c:
            raise HTTPException(404, "Không tìm thấy hợp đồng")
        if not c.output_file_path or not os.path.exists(c.output_file_path):
            raise HTTPException(404, "File chưa được sinh hoặc không còn tồn tại")
        path = c.output_file_path
        name = c.output_file_name or os.path.basename(path)
    finally:
        db.close()

    return FileResponse(
        path=path, filename=name,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )

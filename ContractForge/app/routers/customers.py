"""
Router: /customers — CRUD khách hàng + import Excel.

Routes:
  GET  /customers                → danh sách + lọc + phân trang
  GET  /customers/new            → form tạo mới
  POST /customers/new            → xử lý tạo mới
  GET  /customers/{id}           → chi tiết
  GET  /customers/{id}/edit      → form sửa
  POST /customers/{id}/edit      → xử lý sửa
  POST /customers/{id}/toggle    → bật/tắt is_active
  GET  /customers/import         → trang import Excel
  POST /customers/import         → xử lý upload Excel
  GET  /customers/template.xlsx  → tải file Excel mẫu
"""

import io, os, math, time
from datetime import datetime, date as _date
from fastapi import APIRouter, Request, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.templating import Jinja2Templates
from typing import Optional

from app.database import SessionLocal
from app.paths import TEMPLATES_DIR, CUSTOMER_DOCS_DIR
from app.models.customer_unit import UNIT_TYPES
from app.models.customer_document import CustomerDocument
from app.services.customer_service import (
    CustomerFilters, list_customers, get_customer,
    create_customer, update_customer, toggle_active,
    get_filter_options, import_from_excel, import_grouped_from_excel,
    generate_next_code,
    delete_customer, bulk_delete_customers,
    restore_customer,
    get_sub_units, save_sub_units,
)


def _parse_date(s: str):
    """Parse chuỗi ngày từ form (YYYY-MM-DD hoặc DD/MM/YYYY) → date | None."""
    from app.utils.date_helpers import parse_date_input
    try:
        return parse_date_input(s) if s and s.strip() else None
    except Exception:
        return None

def _extract_sub_units(form) -> list[dict]:
    """Parse đơn vị con từ form data (các field sub_unit_name[], sub_unit_type[]...)."""
    names      = form.getlist("sub_unit_name")
    types      = form.getlist("sub_unit_type")
    addresses  = form.getlist("sub_unit_address")
    notes      = form.getlist("sub_unit_notes")
    ids        = form.getlist("sub_unit_id")
    result = []
    for i, name in enumerate(names):
        if name.strip():
            uid = ids[i] if i < len(ids) else None
            # Checkbox include_in_billing: tên field là sub_unit_billing_{id}
            # Nếu là sub_unit mới (chưa có id), mặc định True
            in_billing = True
            if uid:
                in_billing = f"sub_unit_billing_{uid}" in form
            result.append({
                "id":               uid,
                "name":             name.strip(),
                "unit_type":        types[i].strip() if i < len(types) else "Điểm trạm",
                "address":          addresses[i].strip() if i < len(addresses) else "",
                "notes":            notes[i].strip() if i < len(notes) else "",
                "include_in_billing": in_billing,
            })
    return result



router = APIRouter(prefix="/customers", tags=["customers"])
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# ── Document upload constants ─────────────────────────────────────────────
_DOC_LABELS = {
    "cccd_front":    "CCCD mặt trước (người đại diện)",
    "cccd_back":     "CCCD mặt sau (người đại diện)",
    "establishment": "Quyết định thành lập",
    "other":         "Giấy tờ khác",
}
_FIXED_DOC_TYPES = {"cccd_front", "cccd_back", "establishment"}
_MAX_DOC_BYTES   = 5 * 1024 * 1024  # 5 MB
_DOC_MIME_MAP    = {
    ".pdf":  ("pdf",   b"%PDF"),
    ".jpg":  ("image", b"\xff\xd8\xff"),
    ".jpeg": ("image", b"\xff\xd8\xff"),
    ".png":  ("image", b"\x89PNG"),
}


def _check_doc_upload(file_bytes: bytes) -> tuple[str, str]:
    """Validate magic bytes. Returns (mime_hint, error) — error="" nếu hợp lệ."""
    for ext, (hint, magic) in _DOC_MIME_MAP.items():
        if file_bytes[:len(magic)] == magic:
            return hint, ""
    return "", "File không hợp lệ — chỉ chấp nhận PDF, JPG hoặc PNG"


def _get_doc_or_404(db, customer_id: int, doc_id: int) -> CustomerDocument:
    """Lấy doc, verify IDOR (customer_id phải khớp)."""
    doc = db.query(CustomerDocument).filter(CustomerDocument.id == doc_id).first()
    if not doc or doc.customer_id != customer_id:
        raise HTTPException(404, "Không tìm thấy giấy tờ")
    return doc


def _safe_header_filename(filename: str) -> str:
    """Content-Disposition an toàn với tên file tiếng Việt (RFC 5987)."""
    import urllib.parse
    encoded    = urllib.parse.quote(filename, safe=".-_")
    ascii_name = "".join(
        c if (c.isascii() and c not in r'\/:*?"<>|') else "_"
        for c in filename
    )
    return (
        "attachment; "
        + 'filename="' + ascii_name + '"; '
        + "filename*=UTF-8''" + encoded
    )

PER_PAGE = 25
CUSTOMER_TYPES = ["Trạm y tế", "TTYT", "Đơn vị quản lý", "Bệnh viện", "Khác"]


# ---------------------------------------------------------------------------
# Helper: flash message qua query param
# ---------------------------------------------------------------------------

def _flash_redirect(url: str, msg: str, msg_type: str = "success") -> RedirectResponse:
    sep = "&" if "?" in url else "?"
    return RedirectResponse(f"{url}{sep}msg={msg}&msg_type={msg_type}", status_code=303)


def _extract_flash(request: Request) -> dict:
    return {
        "flash_msg":  request.query_params.get("msg", ""),
        "flash_type": request.query_params.get("msg_type", "success"),
    }


# ---------------------------------------------------------------------------
# LIST
# ---------------------------------------------------------------------------

ALLOWED_PER_PAGE = [20, 40, 60, 80, 100]

@router.get("", response_class=HTMLResponse)
def customer_list(
    request: Request,
    q: str = "",
    ward: str = "",
    customer_type: str = "",
    active_only: int = 1,
    page: int = 1,
    sort_by: str = "ward",
    sort_dir: str = "asc",
    per_page: int = 25,
):
    per_page = per_page if per_page in ALLOWED_PER_PAGE else 25
    db = SessionLocal()
    try:
        filters = CustomerFilters(
            q=q, ward=ward, customer_type=customer_type,
            active_only=bool(active_only), page=page, per_page=per_page,
            sort_by=sort_by, sort_dir=sort_dir,
        )
        rows, total = list_customers(db, filters)
        options = get_filter_options(db)
        total_pages = max(1, math.ceil(total / per_page))
        customers_with_count = [(c, cnt, act or 0, exp or 0) for c, cnt, act, exp in rows]
    finally:
        db.close()

    filter_parts = [ward or "Tất cả địa bàn cấp xã", customer_type or "Tất cả loại đơn vị"]
    if q.strip():
        filter_parts.append(f"Tìm kiếm: {q.strip()}")
    filter_summary_value = " · ".join(filter_parts)
    filter_summary_subtext = "Chỉ lấy khách hàng đang hoạt động" if bool(active_only) else "Hiển thị toàn bộ khách hàng"

    return templates.TemplateResponse(request, "customers/list.html", {
        "page_title": "Khách hàng",
        "customers_with_count": customers_with_count,
        "total": total,
        "filters": filters,
        "options": options,
        "page": page,
        "total_pages": total_pages,
        "per_page": per_page,
        "allowed_per_page": ALLOWED_PER_PAGE,
        "filter_summary_value": filter_summary_value,
        "filter_summary_subtext": filter_summary_subtext,
        "sort_by": sort_by,
        "sort_dir": sort_dir,
        **_extract_flash(request),
    })


# ---------------------------------------------------------------------------
# IMPORT EXCEL — phải đặt TRƯỚC route /{id} để tránh bị bắt nhầm
# ---------------------------------------------------------------------------

@router.get("/import", response_class=HTMLResponse)
def import_form(request: Request):
    return templates.TemplateResponse(request, "customers/import.html", {
        "page_title": "Khách hàng",
        **_extract_flash(request),
    })


@router.post("/import")
async def import_excel(
    request: Request,
    file: UploadFile = File(...),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        return templates.TemplateResponse(request, "customers/import.html", {
            "page_title": "Khách hàng",
            "flash_msg": "Chỉ chấp nhận file .xlsx hoặc .xls",
            "flash_type": "danger",
        })

    file_bytes = await file.read()
    if len(file_bytes) > 5 * 1024 * 1024:  # 5 MB limit
        return templates.TemplateResponse(request, "customers/import.html", {
            "page_title": "Khách hàng",
            "flash_msg": "File quá lớn (giới hạn 5 MB)",
            "flash_type": "danger",
        })

    db = SessionLocal()
    try:
        result = import_from_excel(db, file_bytes, file.filename)
    finally:
        db.close()

    return templates.TemplateResponse(request, "customers/import.html", {
        "page_title": "Khách hàng",
        "import_result": result,
        "filename": file.filename,
    })


@router.get("/import-grouped", response_class=HTMLResponse)
def import_grouped_form(request: Request):
    return templates.TemplateResponse(request, "customers/import.html", {
        "page_title": "Khách hàng",
        "active_tab": "grouped",
        **_extract_flash(request),
    })


@router.post("/import-grouped")
async def import_grouped_excel(
    request: Request,
    file: UploadFile = File(...),
    sheet_name: str = Form(""),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        return templates.TemplateResponse(request, "customers/import.html", {
            "page_title": "Khách hàng",
            "active_tab": "grouped",
            "flash_msg": "Chỉ chấp nhận file .xlsx hoặc .xls",
            "flash_type": "danger",
        })

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        return templates.TemplateResponse(request, "customers/import.html", {
            "page_title": "Khách hàng",
            "active_tab": "grouped",
            "flash_msg": "File quá lớn (giới hạn 10 MB)",
            "flash_type": "danger",
        })

    db = SessionLocal()
    try:
        result = import_grouped_from_excel(
            db, file_bytes, file.filename,
            sheet_name=sheet_name.strip() or None,
        )
    finally:
        db.close()

    return templates.TemplateResponse(request, "customers/import.html", {
        "page_title": "Khách hàng",
        "active_tab": "grouped",
        "grouped_result": result,
        "filename": file.filename,
    })


@router.get("/template.xlsx")
def download_template():
    """Trả về file Excel mẫu để người dùng điền vào rồi import."""
    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách khách hàng"

    headers = [
        "Mã khách hàng", "Tên pháp lý", "Tên ngắn", "Loại đơn vị",
        "Địa chỉ", "Xã / Phường / Đặc khu", "Tỉnh / Thành phố",
        "Mã số thuế", "Mã cơ sở KCB", "Số QĐ thành lập", "Ngày thành lập",
        "Đại diện pháp lý", "Giới tính đại diện", "Chức vụ",
        "Số tài khoản", "Ngân hàng", "Tên thụ hưởng",
        "Đầu mối liên hệ", "Điện thoại", "Email", "Ghi chú",
    ]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1A2238")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Dòng ví dụ
    ws.append([
        "KH-001", "Trạm Y tế xã Đại Phước", "TYT Đại Phước", "Trạm y tế",
        "Ấp Bến Sắn", "Xã Đại Phước", "Đồng Nai",
        "", "", "1234/QĐ-UBND", "01/01/2010",
        "Nguyễn Văn A", "Nam", "Trưởng Trạm",
        "", "", "",
        "", "", "", "",
    ])

    # Column widths (21 cột)
    widths = [14, 35, 20, 15, 25, 22, 16, 14, 16, 20, 14, 22, 12, 16, 18, 25, 25, 20, 14, 22, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mau_import_khachhang.xlsx"},
    )


# ---------------------------------------------------------------------------
# EXPORT REPORT
# ---------------------------------------------------------------------------

@router.get("/export-report.xlsx")
def export_report():
    """Xuất báo cáo tổng hợp khách hàng (4 sheets) dạng Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from sqlalchemy.orm import joinedload
    from fastapi.responses import StreamingResponse
    from app.models.customer import Customer
    from app.models.contract import Contract, ContractStatus
    from datetime import date as _dt_date

    # ── Helpers ───────────────────────────────────────────────────────────────
    TODAY = _dt_date.today()

    STATUS_VI = {
        "Draft":        "Nháp",
        "Generated":    "Đã sinh file",
        "Sent":         "Đã gửi",
        "Signed":       "Đã ký",
        "Active":       "Đang hiệu lực",
        "ExpiringSoon": "Sắp hết hạn",
        "Expired":      "Hết hạn",
        "Terminated":   "Đã thanh lý",
    }

    ACTIVE_STATUSES  = {ContractStatus.Active, ContractStatus.ExpiringSoon}
    SIGNED_STATUSES  = {ContractStatus.Signed, ContractStatus.Active,
                        ContractStatus.ExpiringSoon, ContractStatus.Expired,
                        ContractStatus.Terminated}
    DRAFT_STATUSES   = {ContractStatus.Draft, ContractStatus.Generated, ContractStatus.Sent}
    INACTIVE_STATUSES = {ContractStatus.Expired, ContractStatus.Terminated}

    TIER_COLOR = {
        1: "FFD7D7",  # đỏ nhạt
        2: "FFE8C8",  # cam nhạt
        3: "FFF9C4",  # vàng nhạt
        4: "D7F2D7",  # xanh lá nhạt
        5: "EBEBEB",  # xám nhạt
    }
    TIER_LABEL = {
        1: "🔴 KHẨN",
        2: "🟠 Cần xử lý",
        3: "🟡 Upsell",
        4: "✅ Ổn định",
        5: "⚫ Không HĐ",
    }

    def _fmt_money(v):
        if v is None: return ""
        try:
            n = int(v)
            return f"{n:,}".replace(",", ".")
        except Exception:
            return str(v)

    def _fmt_date(d):
        if d is None: return ""
        try: return d.strftime("%d/%m/%Y")
        except Exception: return str(d)

    def _products_of(contracts):
        """Tập tên sản phẩm ngắn gọn: 'YTCS', 'HSSK'."""
        names = set()
        for c in contracts:
            if c.product and c.product.name:
                n = c.product.name.upper()
                if "YTCS" in n or "Y TẾ CƠ SỞ" in n:
                    names.add("YTCS")
                elif "HSSK" in n or "HỒ SƠ" in n:
                    names.add("HSSK")
                else:
                    names.add(c.product.name[:10])
        return names

    def _classify(customer, all_contracts):
        """Trả về (tier, action_text, days_to_exp, nearest_expiry_contract)."""
        active   = [c for c in all_contracts if c.status in ACTIVE_STATUSES]
        signed   = [c for c in all_contracts if c.status in SIGNED_STATUSES]
        drafts   = [c for c in all_contracts if c.status in DRAFT_STATUSES]
        inactive = [c for c in all_contracts if c.status in INACTIVE_STATUSES]

        units   = customer.total_units
        money_f = lambda c: _fmt_money(c.total_amount) if c.total_amount else "?"

        # Sản phẩm
        active_prods  = _products_of(active)
        all_prods     = _products_of(all_contracts)
        missing_prods = {"YTCS", "HSSK"} - all_prods if all_prods else set()

        # ── Tier 1: HĐ active, hết hạn ≤30 ngày ────────────────────────────
        expiring_soon = [c for c in active if c.end_date and
                         0 <= (c.end_date - TODAY).days <= 30]
        if expiring_soon:
            c = min(expiring_soon, key=lambda x: x.end_date)
            days = (c.end_date - TODAY).days
            prod = list(_products_of([c]))[0] if _products_of([c]) else "HĐ"
            txt = (f"Gia hạn {prod} — còn {days} ngày | "
                   f"{units} cơ sở | {money_f(c)}đ")
            return 1, txt, days, c

        # ── Tier 2a: HĐ active, hết hạn 31–60 ngày ──────────────────────────
        expiring_60 = [c for c in active if c.end_date and
                       31 <= (c.end_date - TODAY).days <= 60]
        if expiring_60:
            c = min(expiring_60, key=lambda x: x.end_date)
            days = (c.end_date - TODAY).days
            prod = list(_products_of([c]))[0] if _products_of([c]) else "HĐ"
            txt = (f"Chuẩn bị gia hạn {prod} — còn {days} ngày | "
                   f"{units} cơ sở | {money_f(c)}đ")
            return 2, txt, days, c

        # ── Tier 2b: Không có HĐ active, có HĐ đã hết hạn/thanh lý ─────────
        if not active and inactive:
            c = max(inactive, key=lambda x: x.end_date or _dt_date.min)
            txt = (f"HĐ {c.contract_number} hết hạn {_fmt_date(c.end_date)} "
                   f"— chưa ký mới | {units} cơ sở")
            days = (c.end_date - TODAY).days if c.end_date else 0
            return 2, txt, days, c

        # ── Tier 2c: HĐ nháp quá lâu (>14 ngày) không có active ─────────────
        if not active and drafts:
            old_drafts = [c for c in drafts if c.created_at and
                          (TODAY - c.created_at.date()).days > 14]
            if old_drafts:
                c = max(old_drafts, key=lambda x: x.created_at)
                age = (TODAY - c.created_at.date()).days
                txt = (f"HĐ nháp {c.contract_number} từ {age} ngày trước "
                       f"— chưa ký | {units} cơ sở")
                return 2, txt, age, c

        # ── Tier 3: Đang active, thiếu sản phẩm ─────────────────────────────
        if active and missing_prods:
            missing_str = " & ".join(sorted(missing_prods))
            has_str     = " & ".join(sorted(active_prods)) or "?"
            c = max(active, key=lambda x: x.end_date or _dt_date.min)
            days = (c.end_date - TODAY).days if c.end_date else 0
            txt = (f"Upsell {missing_str} — đang có {has_str} | "
                   f"{units} cơ sở | còn {days} ngày")
            return 3, txt, days, c

        # ── Tier 4: Đang active, đủ sản phẩm ────────────────────────────────
        if active:
            c = max(active, key=lambda x: x.end_date or _dt_date.min)
            days = (c.end_date - TODAY).days if c.end_date else 0
            prod_str = " & ".join(sorted(active_prods)) or "HĐ"
            active_val = sum(int(x.total_amount) for x in active if x.total_amount)
            txt = (f"{prod_str} hiệu lực đến {_fmt_date(c.end_date)} | "
                   f"{units} cơ sở | {_fmt_money(active_val)}đ")
            return 4, txt, days, c

        # ── Tier 5b: Có HĐ nhưng tất cả đều nháp (và ko quá cũ) ─────────────
        if drafts and not inactive:
            c = max(drafts, key=lambda x: x.created_at or _dt_date.min)
            age = (TODAY - c.created_at.date()).days if c.created_at else 0
            txt = (f"HĐ nháp {c.contract_number} — {STATUS_VI.get(c.status.value, c.status.value)} "
                   f"| {units} cơ sở")
            return 5, txt, 0, c

        # ── Tier 5a: Chưa có HĐ nào ─────────────────────────────────────────
        txt = f"Chưa có hợp đồng | {units} cơ sở — cần tiếp cận"
        return 5, txt, 0, None

    # ── Load data ─────────────────────────────────────────────────────────────
    db = SessionLocal()
    try:
        customers = (
            db.query(Customer)
            .filter(Customer.is_deleted == False)
            .options(
                joinedload(Customer.sub_units),
                joinedload(Customer.contracts).joinedload(Contract.product),
            )
            .order_by(Customer.code)
            .all()
        )
        # Materialize toàn bộ trước khi close session
        cust_data = []
        for c in customers:
            contracts = list(c.contracts)
            tier, action, days, near_c = _classify(c, contracts)
            active_cts = [x for x in contracts if x.status in ACTIVE_STATUSES]
            signed_cts = [x for x in contracts if x.status in SIGNED_STATUSES]
            active_val = sum(int(x.total_amount) for x in active_cts if x.total_amount)
            lifetime_val = sum(int(x.total_amount) for x in contracts if x.total_amount)
            last_sign = max((x.sign_date for x in contracts if x.sign_date),
                            default=None)
            last_contract = (max((x for x in contracts if x.sign_date),
                                 key=lambda x: x.sign_date, default=None)
                             or max(contracts, key=lambda x: x.id, default=None))
            risk_30  = sum(int(x.total_amount) for x in active_cts
                          if x.total_amount and x.end_date
                          and 0 <= (x.end_date - TODAY).days <= 30)
            risk_60  = sum(int(x.total_amount) for x in active_cts
                          if x.total_amount and x.end_date
                          and 0 <= (x.end_date - TODAY).days <= 60)
            prods_active = " & ".join(sorted(_products_of(active_cts))) or "—"
            prods_all    = " & ".join(sorted(_products_of(contracts))) or "—"
            missing      = " & ".join(sorted({"YTCS","HSSK"} - _products_of(contracts))) or "—"
            gender_label = {"Nam": "(Ông)", "Nữ": "(Bà)"}.get(
                c.representative_gender or "", "")
            rep_display  = f"{gender_label} {c.representative_name}".strip() if c.representative_name else ""
            cust_data.append({
                "obj":           c,
                "contracts":     contracts,
                "active_cts":    active_cts,
                "signed_cts":    signed_cts,
                "tier":          tier,
                "action":        action,
                "days":          days,
                "near_c":        near_c,
                "active_val":    active_val,
                "lifetime_val":  lifetime_val,
                "last_sign":     last_sign,
                "last_contract": last_contract,
                "risk_30":       risk_30,
                "risk_60":       risk_60,
                "prods_active":  prods_active,
                "prods_all":     prods_all,
                "missing":       missing,
                "rep_display":   rep_display,
            })
    finally:
        db.close()

    # ── Workbook helpers ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    HDR_FILL   = PatternFill("solid", fgColor="1A2238")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DATA_ALIGN = Alignment(vertical="center", wrap_text=False)
    WRAP_ALIGN = Alignment(vertical="center", wrap_text=True)
    MONEY_FMT  = '#,##0'
    DATE_FMT   = 'DD/MM/YYYY'
    INACTIVE_FONT = Font(color="999999", italic=True, size=10)
    THIN = Side(style='thin', color="CCCCCC")
    THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _make_header(ws, cols, row_h=30):
        ws.append([c[0] for c in cols])
        ws.row_dimensions[1].height = row_h
        for i, cell in enumerate(ws[1]):
            cell.font  = HDR_FONT
            cell.fill  = HDR_FILL
            cell.alignment = HDR_ALIGN
            if i < len(cols):
                ws.column_dimensions[get_column_letter(i+1)].width = cols[i][1]

    def _set_row(ws, row_idx, values, cols, fill=None, font=None, num_fmts=None):
        ws.append(values)
        for i, cell in enumerate(ws[row_idx]):
            cell.alignment = DATA_ALIGN
            if fill: cell.fill = fill
            if font: cell.font = font
            else:    cell.font = Font(size=10)
            cell.border = THIN_BORDER
            if num_fmts and i < len(num_fmts) and num_fmts[i]:
                cell.number_format = num_fmts[i]

    # ── SHEET 1: Ưu tiên hành động ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ưu tiên hành động"

    S1_COLS = [
        ("STT",                      5),
        ("Tier",                    13),
        ("Mã KH",                   10),
        ("Tên ngắn",                22),
        ("Loại đơn vị",             14),
        ("Hành động gợi ý",         48),
        ("Ngày hết hạn HĐ gần nhất",16),
        ("Còn (ngày)",               10),
        ("Số HĐ gần nhất",          22),
        ("Sản phẩm active",         14),
        ("Còn thiếu",               12),
        ("Số cơ sở",                 9),
        ("Giá trị active (VNĐ)",    18),
        ("Nguy cơ mất 30 ngày",     18),
        ("Nguy cơ mất 60 ngày",     18),
    ]
    _make_header(ws1, S1_COLS)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(S1_COLS))}1"

    sorted_data = sorted(cust_data, key=lambda x: (x["tier"], x["days"] if x["tier"] <= 2 else -x["days"]))

    for idx, d in enumerate(sorted_data, 1):
        c   = d["obj"]
        nc  = d["near_c"]
        tier = d["tier"]
        fill = PatternFill("solid", fgColor=TIER_COLOR[tier])
        font = INACTIVE_FONT if not c.is_active else Font(size=10)
        vals = [
            idx,
            TIER_LABEL[tier],
            c.code or "",
            c.short_name or c.legal_name or "",
            c.customer_type or "",
            d["action"],
            nc.end_date if nc and nc.end_date else None,
            d["days"] if nc else None,
            nc.contract_number if nc else "",
            d["prods_active"],
            d["missing"],
            c.total_units,
            d["active_val"] or None,
            d["risk_30"] or None,
            d["risk_60"] or None,
        ]
        ws1.append(vals)
        row_idx = idx + 1
        ws1.row_dimensions[row_idx].height = 20
        for ci, cell in enumerate(ws1[row_idx]):
            cell.fill   = fill
            cell.font   = font
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN if ci == 5 else DATA_ALIGN
            if ci == 6 and cell.value:  # ngày hết hạn
                cell.number_format = DATE_FMT
            if ci in (12, 13, 14) and cell.value:
                cell.number_format = MONEY_FMT

    # ── SHEET 2: Hồ sơ đầy đủ ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Hồ sơ đầy đủ")
    S2_COLS = [
        ("STT",               5), ("Mã KH",           10), ("Tên pháp lý",      32),
        ("Tên ngắn",         20), ("Loại đơn vị",     14), ("Xã/Phường",        18),
        ("Tỉnh/TP",          16), ("Trạng thái KH",   12), ("Số cơ sở",          9),
        ("Đại diện pháp lý", 22), ("Chức vụ",         14), ("SĐT",              14),
        ("Email",            22), ("Mã số thuế",      14), ("Mã QHNS",          12),
        ("Mã KCB BHYT",      14), ("Số QĐ thành lập", 18), ("Ngày thành lập",   14),
        ("Ngân hàng",        24), ("Số TK",           18),
        ("Tổng HĐ đã ký",   12), ("HĐ đang hiệu lực",14),
        ("Sản phẩm đang có", 16), ("Sản phẩm còn thiếu",16),
        ("Giá trị active",  18), ("Giá trị lifetime", 18),
        ("Ngày ký gần nhất", 14), ("Số HĐ gần nhất",  22),
        ("Ghi chú",         24),
    ]
    _make_header(ws2, S2_COLS)
    ws2.freeze_panes = "C2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(S2_COLS))}1"

    for idx, d in enumerate(cust_data, 1):
        c  = d["obj"]
        font = INACTIVE_FONT if not c.is_active else Font(size=10)
        lc = d["last_contract"]
        vals = [
            idx, c.code or "", c.legal_name or "", c.short_name or "",
            c.customer_type or "", c.ward or "", c.province or "",
            "Hoạt động" if c.is_active else "Tạm ngưng",
            c.total_units,
            d["rep_display"], c.representative_title or "",
            c.phone or "", c.email or "",
            c.tax_code or "", c.budget_relation_code or "",
            c.kcb_code or "", c.establishment_decision or "",
            c.establishment_date if c.establishment_date else None,
            c.bank_name or "", c.bank_account or "",
            len(d["signed_cts"]), len(d["active_cts"]),
            d["prods_all"], d["missing"],
            d["active_val"] or None, d["lifetime_val"] or None,
            d["last_sign"] if d["last_sign"] else None,
            lc.contract_number if lc else "",
            c.notes or "",
        ]
        ws2.append(vals)
        row_idx = idx + 1
        ws2.row_dimensions[row_idx].height = 18
        for ci, cell in enumerate(ws2[row_idx]):
            cell.font   = font
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGN
            if ci == 17 and cell.value:  # ngày thành lập
                cell.number_format = DATE_FMT
            if ci == 26 and cell.value:  # ngày ký gần nhất
                cell.number_format = DATE_FMT
            if ci in (24, 25) and cell.value:
                cell.number_format = MONEY_FMT

    # ── SHEET 3: Lịch sử hợp đồng ────────────────────────────────────────────
    ws3 = wb.create_sheet("Lịch sử hợp đồng")
    S3_COLS = [
        ("STT",             5), ("Mã KH",         10), ("Tên KH",          22),
        ("Số hợp đồng",    24), ("Loại SP",        12), ("Trạng thái",      14),
        ("Ngày ký",        12), ("Hiệu lực từ",   12), ("Hết hạn",         12),
        ("Số tháng",        9), ("Số cơ sở",        9),
        ("Đơn giá (trước VAT)",17), ("Đơn giá (sau VAT)",16),
        ("Tổng tiền (VNĐ)", 18),
    ]
    _make_header(ws3, S3_COLS)
    ws3.freeze_panes = "D2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(S3_COLS))}1"

    ct_idx = 0
    for d in cust_data:
        c = d["obj"]
        cts = sorted(d["contracts"], key=lambda x: (x.sign_date or _dt_date.min), reverse=True)
        for ct in cts:
            ct_idx += 1
            prod_name = ct.product.name if ct.product else ""
            prod_short = ""
            if prod_name:
                pn = prod_name.upper()
                if "YTCS" in pn or "Y TẾ CƠ SỞ" in pn: prod_short = "YTCS"
                elif "HSSK" in pn or "HỒ SƠ" in pn: prod_short = "HSSK"
                else: prod_short = prod_name[:10]
            vals = [
                ct_idx, c.code or "", c.short_name or c.legal_name or "",
                ct.contract_number or "",
                prod_short,
                STATUS_VI.get(ct.status.value if hasattr(ct.status,'value') else str(ct.status), ""),
                ct.sign_date, ct.start_date, ct.end_date,
                ct.month_count,
                ct.unit_count,
                ct.unit_price_pre_vat or None,
                ct.unit_price_vat or None,
                int(ct.total_amount) if ct.total_amount else None,
            ]
            ws3.append(vals)
            row_idx = ct_idx + 1
            ws3.row_dimensions[row_idx].height = 18
            for ci, cell in enumerate(ws3[row_idx]):
                cell.font   = Font(size=10)
                cell.border = THIN_BORDER
                cell.alignment = DATA_ALIGN
                if ci in (6, 7, 8) and cell.value:
                    cell.number_format = DATE_FMT
                if ci in (11, 12, 13) and cell.value:
                    cell.number_format = MONEY_FMT

    # ── SHEET 4: Tóm tắt ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Tóm tắt")
    ws4.column_dimensions["A"].width = 34
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 14

    TITLE_FONT  = Font(bold=True, size=13, color="1A2238")
    LABEL_FONT  = Font(size=10)
    VALUE_FONT  = Font(bold=True, size=11, color="1A2238")
    SEC_FILL    = PatternFill("solid", fgColor="E8EDF5")
    SEC_FONT    = Font(bold=True, size=10, color="1A2238")

    def _s4_title(row, text):
        c = ws4.cell(row=row, column=1, value=text)
        c.font = TITLE_FONT

    def _s4_section(row, text):
        for col in range(1, 4):
            ws4.cell(row=row, column=col).fill = SEC_FILL
        c = ws4.cell(row=row, column=1, value=text)
        c.font = SEC_FONT
        ws4.row_dimensions[row].height = 18

    def _s4_row(row, label, value, fmt=None):
        ws4.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell = ws4.cell(row=row, column=2, value=value)
        cell.font  = VALUE_FONT
        if fmt: cell.number_format = fmt
        ws4.row_dimensions[row].height = 16

    active_kh   = sum(1 for d in cust_data if d["obj"].is_active)
    inactive_kh = len(cust_data) - active_kh
    no_contract = sum(1 for d in cust_data if not d["contracts"])
    upsell_cnt  = sum(1 for d in cust_data if d["tier"] == 3)
    total_active_rev = sum(d["active_val"] for d in cust_data)
    total_life_rev   = sum(d["lifetime_val"] for d in cust_data)
    risk_30_total    = sum(d["risk_30"] for d in cust_data)
    risk_60_total    = sum(d["risk_60"] for d in cust_data)
    tier_counts = {t: sum(1 for d in cust_data if d["tier"] == t) for t in range(1, 6)}

    r = 1
    _s4_title(r, f"BÁO CÁO TỔNG HỢP KHÁCH HÀNG — {TODAY.strftime('%d/%m/%Y')}"); r += 2

    _s4_section(r, "TỔNG QUAN KHÁCH HÀNG"); r += 1
    _s4_row(r, "Tổng số khách hàng", len(cust_data)); r += 1
    _s4_row(r, "Đang hoạt động", active_kh); r += 1
    _s4_row(r, "Tạm ngưng", inactive_kh); r += 1
    _s4_row(r, "Chưa có hợp đồng nào", no_contract); r += 2

    _s4_section(r, "PHÂN LOẠI THEO ƯU TIÊN"); r += 1
    for t in range(1, 6):
        _s4_row(r, f"{TIER_LABEL[t]}", tier_counts[t]); r += 1
    r += 1

    _s4_section(r, "DOANH THU & RỦI RO"); r += 1
    _s4_row(r, "Tổng giá trị HĐ đang hiệu lực (VNĐ)", total_active_rev, MONEY_FMT); r += 1
    _s4_row(r, "Giá trị có nguy cơ mất trong 30 ngày", risk_30_total, MONEY_FMT); r += 1
    _s4_row(r, "Giá trị có nguy cơ mất trong 60 ngày", risk_60_total, MONEY_FMT); r += 1
    _s4_row(r, "Tổng giá trị lifetime (tất cả HĐ)", total_life_rev, MONEY_FMT); r += 2

    _s4_section(r, "CƠ HỘI"); r += 1
    _s4_row(r, "KH có thể upsell thêm sản phẩm", upsell_cnt); r += 1
    ytcs_only = sum(1 for d in cust_data
                    if _products_of(d["active_cts"]) == {"YTCS"})
    hssk_only  = sum(1 for d in cust_data
                     if _products_of(d["active_cts"]) == {"HSSK"})
    both       = sum(1 for d in cust_data
                     if len(_products_of(d["active_cts"])) >= 2)
    _s4_row(r, "  Đang có YTCS (chưa có HSSK)", ytcs_only); r += 1
    _s4_row(r, "  Đang có HSSK (chưa có YTCS)", hssk_only); r += 1
    _s4_row(r, "  Có cả YTCS + HSSK", both); r += 2

    _s4_section(r, "TOP 5 THEO GIÁ TRỊ HĐ ACTIVE"); r += 1
    top5 = sorted(cust_data, key=lambda x: x["active_val"], reverse=True)[:5]
    for d in top5:
        c = d["obj"]
        label = f"{c.code} — {c.short_name or c.legal_name}"
        _s4_row(r, label, d["active_val"], MONEY_FMT); r += 1

    # ── Save ──────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"BaoCao_KhachHang_{TODAY.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _safe_header_filename(fname)},
    )

@router.get("/new", response_class=HTMLResponse)
def new_form(request: Request):
    db = SessionLocal()
    try:
        next_code = generate_next_code(db)
    finally:
        db.close()

    return templates.TemplateResponse(request, "customers/form.html", {
        "page_title": "Khách hàng",
        "form_title": "Thêm khách hàng mới",
        "unit_types": UNIT_TYPES,
        "sub_units": [],
        "action": "/customers/new",
        "customer": None,
        "next_code": next_code,
        "customer_types": CUSTOMER_TYPES,
        "errors": [],
    })


@router.post("/new")
async def new_submit(
    request: Request,
    code: str = Form(""),
    legal_name: str = Form(...),
    short_name: str = Form(""),
    customer_type: str = Form("Trạm y tế"),
    address: str = Form(""),
    ward: str = Form(""),
    province: str = Form("Đồng Nai"),
    tax_code: str = Form(""),
    budget_relation_code: str = Form(""),
    kcb_code: str = Form(""),
    establishment_decision: str = Form(""),
    establishment_date: str = Form(""),
    representative_name: str = Form(""),
    representative_gender: str = Form(""),
    representative_title: str = Form(""),
    bank_account: str = Form(""),
    bank_name: str = Form(""),
    beneficiary_name: str = Form(""),
    contact_name: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    notes: str = Form(""),
):
    data = {
        "code": code, "legal_name": legal_name, "short_name": short_name,
        "customer_type": customer_type,
        "address": address, "ward": ward, "province": province,
        "tax_code": tax_code, "budget_relation_code": budget_relation_code or None,
        "kcb_code": kcb_code or None,
        "establishment_decision": establishment_decision or None,
        "establishment_date": _parse_date(establishment_date),
        "representative_name": representative_name,
        "representative_gender": representative_gender or None,
        "representative_title": representative_title,
        "bank_account": bank_account, "bank_name": bank_name,
        "beneficiary_name": beneficiary_name,
        "contact_name": contact_name, "phone": phone, "email": email,
        "notes": notes, "is_active": True,
    }
    db = SessionLocal()
    try:
        customer = create_customer(db, data)
        # bill_main_station: đọc đúng checkbox (form tạo mới đã tích sẵn mặc định)
        raw_form = await request.form()
        customer.bill_main_station = "bill_main_station" in raw_form
        sub_units_data = _extract_sub_units(raw_form)
        save_sub_units(db, customer.id, sub_units_data)
        # Save ids before session closes
        new_id   = customer.id
        new_code = customer.code
    except ValueError as e:
        db.close()
        next_code = code or ""
        return templates.TemplateResponse(request, "customers/form.html", {
            "page_title": "Khách hàng",
            "form_title": "Thêm khách hàng mới",
            "action": "/customers/new",
            "customer": None,
            "next_code": next_code,
            "customer_types": CUSTOMER_TYPES,
            "unit_types": UNIT_TYPES,
            "sub_units": [],
            "errors": [str(e)],
            "prefill": data,
        }, status_code=422)
    finally:
        db.close()

    return _flash_redirect(f"/customers/{new_id}", f"Đã tạo khách hàng {new_code}", "success")


# ---------------------------------------------------------------------------
# DETAIL
# ---------------------------------------------------------------------------


# ─── BULK DELETE ─────────────────────────────────────────────────────────────

@router.post("/bulk-delete")
async def bulk_delete(request: Request):
    form = await request.form()
    ids_raw = form.getlist("selected_ids")
    try:
        ids = [int(i) for i in ids_raw if str(i).strip().isdigit()]
    except ValueError:
        ids = []
    if not ids:
        return _flash_redirect("/customers", "Chưa chọn khách hàng nào", "warning")
    db = SessionLocal()
    try:
        n = bulk_delete_customers(db, ids)
    finally:
        db.close()
    return _flash_redirect("/customers",
        f"Đã chuyển {n} khách hàng vào thùng rác — khôi phục được trong Cài đặt",
        "warning")


@router.post("/{customer_id}/delete")
def delete_one(customer_id: int):
    db = SessionLocal()
    try:
        c = delete_customer(db, customer_id)
    finally:
        db.close()
    if c:
        name = c.short_name or c.legal_name
        return _flash_redirect("/customers",
            f"Đã chuyển '{name}' vào thùng rác — khôi phục được trong Cài đặt", "warning")
    return _flash_redirect("/customers", "Không tìm thấy khách hàng", "danger")


@router.get("/{customer_id}", response_class=HTMLResponse)
def customer_detail(request: Request, customer_id: int):
    db = SessionLocal()
    try:
        customer = get_customer(db, customer_id)
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        if customer.is_deleted:
            return _flash_redirect("/customers",
                "Khách hàng này đã bị xóa — khôi phục trong Cài đặt nếu cần", "warning")
        contract_count = len(customer.contracts)
        sub_units = get_sub_units(db, customer_id)
        docs = list(db.query(CustomerDocument)
                    .filter(CustomerDocument.customer_id == customer_id)
                    .order_by(CustomerDocument.doc_type, CustomerDocument.id)
                    .all())
    finally:
        db.close()

    return templates.TemplateResponse(request, "customers/detail.html", {
        "page_title": "Khách hàng",
        "customer": customer,
        "contract_count": contract_count,
        "sub_units": sub_units,
        "docs": docs,
        "doc_labels": _DOC_LABELS,
        **_extract_flash(request),
    })


# ---------------------------------------------------------------------------
# EDIT
# ---------------------------------------------------------------------------

@router.get("/{customer_id}/edit", response_class=HTMLResponse)
def edit_form(request: Request, customer_id: int):
    db = SessionLocal()
    try:
        customer = get_customer(db, customer_id)
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        sub_units = get_sub_units(db, customer_id)
        _code = customer.code
        _title = customer.short_name or customer.legal_name
    finally:
        db.close()

    return templates.TemplateResponse(request, "customers/form.html", {
        "page_title": "Khách hàng",
        "form_title": f"Sửa — {_title}",
        "action": f"/customers/{customer_id}/edit",
        "customer": customer,
        "next_code": _code,
        "customer_types": CUSTOMER_TYPES,
        "unit_types": UNIT_TYPES,
        "sub_units": sub_units,
        "errors": [],
    })


@router.post("/{customer_id}/edit")
async def edit_submit(
    request: Request,
    customer_id: int,
    code: str = Form(""),
    legal_name: str = Form(...),
    short_name: str = Form(""),
    customer_type: str = Form("Trạm y tế"),
    address: str = Form(""),
    ward: str = Form(""),
    province: str = Form("Đồng Nai"),
    tax_code: str = Form(""),
    budget_relation_code: str = Form(""),
    kcb_code: str = Form(""),
    establishment_decision: str = Form(""),
    establishment_date: str = Form(""),
    representative_name: str = Form(""),
    representative_gender: str = Form(""),
    representative_title: str = Form(""),
    bank_account: str = Form(""),
    bank_name: str = Form(""),
    beneficiary_name: str = Form(""),
    contact_name: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    notes: str = Form(""),
):
    # Nếu form không gửi code (hoặc để trống), giữ nguyên code cũ
    # bằng cách không đưa vào data dict → update_customer sẽ không động vào nó.
    data = {
        "legal_name": legal_name, "short_name": short_name,
        "customer_type": customer_type,
        "address": address, "ward": ward, "province": province,
        "tax_code": tax_code, "budget_relation_code": budget_relation_code or None,
        "kcb_code": kcb_code or None,
        "establishment_decision": establishment_decision or None,
        "establishment_date": _parse_date(establishment_date),
        "representative_name": representative_name,
        "representative_gender": representative_gender or None,
        "representative_title": representative_title,
        "bank_account": bank_account, "bank_name": bank_name,
        "beneficiary_name": beneficiary_name,
        "contact_name": contact_name, "phone": phone, "email": email,
        "notes": notes,
    }
    # Chỉ thêm code vào data nếu người dùng thực sự điền vào
    if code.strip():
        data["code"] = code.strip()

    db = SessionLocal()
    try:
        customer = update_customer(db, customer_id, data)
        raw_form = await request.form()
        # bill_main_station: checkbox có mặt → True, vắng mặt → False
        customer.bill_main_station = "bill_main_station" in raw_form
        sub_units_data = _extract_sub_units(raw_form)
        save_sub_units(db, customer.id, sub_units_data)
        db.commit()
    except (ValueError, LookupError) as e:
        db2 = SessionLocal()
        orig = get_customer(db2, customer_id)
        db2.close()
        db.close()
        return templates.TemplateResponse(request, "customers/form.html", {
            "page_title": "Khách hàng",
            "form_title": "Sửa khách hàng",
            "action": f"/customers/{customer_id}/edit",
            "customer": orig,
            "next_code": code,
            "customer_types": CUSTOMER_TYPES,
            "errors": [str(e)],
            "prefill": data,
        }, status_code=422)
    finally:
        db.close()

    return _flash_redirect(f"/customers/{customer_id}", "Đã cập nhật thành công", "success")


# ---------------------------------------------------------------------------
# TOGGLE ACTIVE
# ---------------------------------------------------------------------------

@router.post("/{customer_id}/toggle")
def toggle(request: Request, customer_id: int):
    db = SessionLocal()
    try:
        customer = toggle_active(db, customer_id)
    except LookupError as e:
        raise HTTPException(status_code=404, detail=str(e))
    finally:
        db.close()

    status_text = "kích hoạt" if customer.is_active else "tạm ngưng"
    return _flash_redirect(
        f"/customers/{customer_id}",
        f"Đã {status_text} khách hàng {customer.code}",
        "success" if customer.is_active else "warning",
    )


# ---------------------------------------------------------------------------
# BILLING CONFIG — chọn cơ sở tính phí
# ---------------------------------------------------------------------------

@router.post("/{customer_id}/billing")
async def save_billing(request: Request, customer_id: int):
    """Lưu cấu hình cơ sở tính phí: bill_main_station + include_in_billing cho từng sub_unit."""
    raw_form = await request.form()
    db = SessionLocal()
    try:
        from app.models.customer import Customer
        from app.models.customer_unit import CustomerUnit

        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")

        # Trạm chính: checkbox tên "bill_main" có mặt → True
        customer.bill_main_station = "bill_main" in raw_form

        # Các sub_unit: checkbox tên "unit_{id}" có mặt → include_in_billing = True
        units = db.query(CustomerUnit).filter(CustomerUnit.customer_id == customer_id).all()
        for u in units:
            u.include_in_billing = f"unit_{u.id}" in raw_form

        db.commit()
    finally:
        db.close()

    return _flash_redirect(
        f"/customers/{customer_id}",
        "Đã lưu cấu hình cơ sở tính phí",
        "success",
    )

# ---------------------------------------------------------------------------
# EXPORT INFO — xuất file Markdown tổng hợp thông tin khách hàng
# ---------------------------------------------------------------------------

def _build_info_md(customer, sub_units: list, contracts: list = None) -> str:
    """Build nội dung file .md tổng hợp thông tin khách hàng.
    Dùng chung cho export_info và download_bundle.
    contracts: list of dict với keys number/status/product/sign_date/end_date (pre-materialized)."""
    if contracts is None:
        contracts = []
    from datetime import datetime as _dt

    def _v(val):
        return str(val).strip() if val else ""

    name    = customer.short_name or customer.legal_name
    now_str = _dt.now().strftime("%d/%m/%Y %H:%M")

    lines = [
        f"# {name} — Thông tin khách hàng",
        f"*Xuất lúc: {now_str}*",
        "",
        "---",
        "",
        "## Thông tin cơ bản",
        f"- **Mã KH:** {_v(customer.code)}",
        f"- **Tên pháp lý:** {_v(customer.legal_name)}",
        f"- **Tên ngắn:** {_v(customer.short_name)}",
        f"- **Loại đơn vị:** {_v(customer.customer_type)}",
        "",
        "## Địa chỉ",
        f"- **Địa chỉ:** {_v(customer.address)}",
        f"- **Xã / Phường:** {_v(customer.ward)}",
        f"- **Tỉnh / TP:** {_v(customer.province)}",
        f"- **Địa chỉ đầy đủ:** {customer.full_address()}",
        "",
        "## Pháp lý & Mã số",
        f"- **Mã số thuế:** {_v(customer.tax_code)}",
        f"- **Mã QHNS:** {_v(customer.budget_relation_code)}",
        f"- **Mã cơ sở KCB BHYT:** {_v(customer.kcb_code)}",
        f"- **Số QĐ thành lập:** {_v(customer.establishment_decision)}",
        f"- **Ngày thành lập:** {customer.establishment_date.strftime('%d/%m/%Y') if customer.establishment_date else ''}",
        f"- **Đại diện pháp lý:** {_v(customer.representative_name)}",
        f"- **Giới tính đại diện:** {'Ông' if customer.representative_gender == 'Nam' else ('Bà' if customer.representative_gender == 'Nữ' else '')}",
        f"- **Chức vụ:** {_v(customer.representative_title)}",
        "",
        "## Ngân hàng",
        f"- **Số tài khoản:** {_v(customer.bank_account)}",
        f"- **Ngân hàng:** {_v(customer.bank_name)}",
        f"- **Tên thụ hưởng:** {_v(customer.beneficiary_name)}",
        "",
        "## Liên hệ",
        f"- **Đầu mối liên hệ:** {_v(customer.contact_name)}",
        f"- **Điện thoại:** {_v(customer.phone)}",
        f"- **Email:** {_v(customer.email)}",
        "",
        "## Cơ sở tính phí",
        f"- **Tổng số cơ sở:** {customer.total_units}",
    ]
    if customer.bill_main_station:
        lines.append(
            f"  - {_v(customer.short_name or customer.legal_name)} *(trạm chính)*"
        )
    for u in sub_units:
        if u.is_active and u.include_in_billing:
            lines.append(f"  - {_v(u.name)} *({_v(u.unit_type)})*")

    # ── Hợp đồng (F1) ────────────────────────────────────────────────────────
    if contracts:
        _STATUS_VI = {
            "Draft":        "Nháp",
            "Generated":    "Đã sinh file",
            "Sent":         "Đã gửi",
            "Signed":       "Đã ký",
            "Active":       "Đang hiệu lực",
            "ExpiringSoon": "Sắp hết hạn",
            "Expired":      "Hết hạn",
            "Terminated":   "Đã thanh lý",
        }
        lines += ["", f"## Hợp đồng ({len(contracts)} hợp đồng)"]
        for ct in contracts:
            product_name = _v(ct.get("product"))
            status_label = _STATUS_VI.get(ct.get("status", ""), ct.get("status", ""))
            sign_str  = ct["sign_date"].strftime("%d/%m/%Y") if ct.get("sign_date") else "—"
            end_str   = ct["end_date"].strftime("%d/%m/%Y") if ct.get("end_date") else "—"
            prod_part = f" · {product_name}" if product_name else ""
            lines.append(
                f"- **{_v(ct.get('number'))}**{prod_part} · {status_label} · {sign_str} → {end_str}"
            )

    if customer.notes:
        lines += ["", "## Ghi chú", _v(customer.notes)]

    return "\n".join(lines) + "\n"


@router.get("/{customer_id}/export-info")
def export_info(customer_id: int):
    """Xuất file .md tổng hợp toàn bộ thông tin khách hàng."""
    from fastapi.responses import Response
    from app.models.contract import Contract

    db = SessionLocal()
    try:
        customer = get_customer(db, customer_id)
        if not customer or customer.is_deleted:
            raise HTTPException(404, "Không tìm thấy khách hàng")
        sub_units = list(get_sub_units(db, customer_id))
        # Materialize contracts trước db.close() để tránh DetachedInstanceError
        raw_contracts = (
            db.query(Contract)
            .filter(Contract.customer_id == customer_id)
            .order_by(Contract.sign_date.desc().nullslast(), Contract.id.desc())
            .all()
        )
        contracts_info = [
            {
                "number":    ct.contract_number,
                "status":    ct.status.value if hasattr(ct.status, "value") else str(ct.status),
                "product":   ct.product.name if ct.product else "",
                "sign_date": ct.sign_date,
                "end_date":  ct.end_date,
            }
            for ct in raw_contracts
        ]
    finally:
        db.close()

    name     = customer.short_name or customer.legal_name
    filename = (f"{customer.code}_{name}.md"
                .replace(" ", "_").replace("/", "-"))
    content  = _build_info_md(customer, sub_units, contracts_info)

    return Response(
        content=content.encode("utf-8"),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": _safe_header_filename(filename)},
    )


@router.get("/{customer_id}/download-bundle")
def download_bundle(customer_id: int):
    """Đóng gói toàn bộ thông tin + file đính kèm thành ZIP và tải về.
    ZIP bao gồm: thong_tin.md + tất cả file ảnh/PDF đã upload."""
    import zipfile, re as _re
    from fastapi.responses import Response

    # ── Label chuẩn cho fixed types ──────────────────────────────────────────
    _ZIP_NAMES = {
        "cccd_front":    "CCCD_mat_truoc",
        "cccd_back":     "CCCD_mat_sau",
        "establishment": "Quyet_dinh_thanh_lap",
    }

    db = SessionLocal()
    try:
        customer = get_customer(db, customer_id)
        if not customer or customer.is_deleted:
            raise HTTPException(404, "Không tìm thấy khách hàng")
        sub_units = list(get_sub_units(db, customer_id))
        docs = list(
            db.query(CustomerDocument)
            .filter(CustomerDocument.customer_id == customer_id)
            .order_by(CustomerDocument.doc_type, CustomerDocument.id)
            .all()
        )
        # Materialize contracts trước db.close()
        from app.models.contract import Contract
        raw_contracts = (
            db.query(Contract)
            .filter(Contract.customer_id == customer_id)
            .order_by(Contract.sign_date.desc().nullslast(), Contract.id.desc())
            .all()
        )
        contracts_info = [
            {
                "number":    ct.contract_number,
                "status":    ct.status.value if hasattr(ct.status, "value") else str(ct.status),
                "product":   ct.product.name if ct.product else "",
                "sign_date": ct.sign_date,
                "end_date":  ct.end_date,
            }
            for ct in raw_contracts
        ]
    finally:
        db.close()

    # ── Build ZIP in memory ───────────────────────────────────────────────────
    buf = io.BytesIO()
    other_idx = 0  # đếm riêng cho "other" để prefix số thứ tự

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as zf:

        # 1. File thong_tin.md — luôn có
        md_content = _build_info_md(customer, sub_units, contracts_info)
        zi_md = zipfile.ZipInfo("thong_tin.md")
        zi_md.flag_bits |= 0x800  # UTF-8 filename
        zf.writestr(zi_md, md_content.encode("utf-8"))

        # 2. Các file đính kèm
        for doc in docs:
            if not os.path.exists(doc.file_path):
                continue  # file bị mất trên disk → bỏ qua, không crash

            ext = os.path.splitext(doc.file_path)[1].lower()

            if doc.doc_type in _ZIP_NAMES:
                zip_name = f"{_ZIP_NAMES[doc.doc_type]}{ext}"
            else:
                # other type: prefix số thứ tự, dùng label làm tên
                other_idx += 1
                label_raw = doc.label or "Giay_to_khac"
                # Xóa ký tự không hợp lệ trong tên file (giữ Unicode)
                label_safe = _re.sub(r'[\\/:*?"<>|]', "_", label_raw).strip()
                zip_name = f"{other_idx:02d}_{label_safe}{ext}"

            # Đảm bảo UTF-8 cho tên file có tiếng Việt
            zi = zipfile.ZipInfo(zip_name)
            zi.flag_bits |= 0x800
            with open(doc.file_path, "rb") as fh:
                zf.writestr(zi, fh.read())

    # ── Tên file ZIP ─────────────────────────────────────────────────────────
    name        = customer.short_name or customer.legal_name
    zip_filename = (f"{customer.code}_{name}.zip"
                    .replace(" ", "_").replace("/", "-"))

    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": _safe_header_filename(zip_filename)},
    )


# ---------------------------------------------------------------------------
# RESTORE FROM TRASH
# ---------------------------------------------------------------------------

@router.post("/{customer_id}/restore")
def restore(customer_id: int):
    db = SessionLocal()
    try:
        c = restore_customer(db, customer_id)
    except LookupError as e:
        return _flash_redirect("/settings", str(e), "danger")
    finally:
        db.close()
    name = c.short_name or c.legal_name
    return _flash_redirect("/settings", f"Đã khôi phục '{name}'", "success")


# ---------------------------------------------------------------------------
# CUSTOMER DOCUMENTS
# ---------------------------------------------------------------------------

@router.post("/{customer_id}/docs/upload")
async def upload_doc(customer_id: int, doc_file: UploadFile = File(...),
                     doc_type: str = Form(...), label: str = Form("")):
    """Upload giấy tờ khách hàng. Fixed types: upsert 1 file. Other: append."""
    # Validate doc_type
    if doc_type not in _DOC_LABELS:
        return _flash_redirect(f"/customers/{customer_id}",
                               "Loại giấy tờ không hợp lệ", "danger")
    if doc_type == "other" and not label.strip():
        return _flash_redirect(f"/customers/{customer_id}",
                               "Vui lòng nhập tên giấy tờ", "danger")

    file_bytes = await doc_file.read()
    if len(file_bytes) > _MAX_DOC_BYTES:
        return _flash_redirect(f"/customers/{customer_id}",
            f"File quá lớn — tối đa 5 MB (hiện tại: {len(file_bytes)//1024} KB)", "danger")

    mime_hint, err = _check_doc_upload(file_bytes)
    if err:
        return _flash_redirect(f"/customers/{customer_id}", err, "danger")

    # Xác định extension từ filename gốc
    raw_ext = os.path.splitext(doc_file.filename or "")[1].lower()
    ext = raw_ext if raw_ext in _DOC_MIME_MAP else (".pdf" if mime_hint == "pdf" else ".jpg")

    cust_dir = CUSTOMER_DOCS_DIR / str(customer_id)
    cust_dir.mkdir(parents=True, exist_ok=True)

    if doc_type in _FIXED_DOC_TYPES:
        dest_name = f"{customer_id}_{doc_type}{ext}"
    else:
        dest_name = f"{customer_id}_other_{int(time.time() * 1000)}{ext}"
    dest_path = cust_dir / dest_name

    # Ghi file mới TRƯỚC, xóa file cũ sau (tránh data loss nếu write fail)
    try:
        dest_path.write_bytes(file_bytes)
    except OSError as e:
        return _flash_redirect(f"/customers/{customer_id}",
                               f"Không thể lưu file: {e}", "danger")

    db = SessionLocal()
    old_file_to_delete = None  # file cũ sẽ xóa SAU khi commit thành công
    try:
        if doc_type in _FIXED_DOC_TYPES:
            existing = db.query(CustomerDocument).filter_by(
                customer_id=customer_id, doc_type=doc_type).first()
            if existing:
                # Ghi nhớ path file cũ, chỉ xóa sau khi commit OK
                if existing.file_path != str(dest_path):
                    old_file_to_delete = existing.file_path
                existing.file_path   = str(dest_path)
                existing.file_name   = doc_file.filename or dest_name
                existing.mime_hint   = mime_hint
                existing.uploaded_at = datetime.utcnow()
            else:
                db.add(CustomerDocument(
                    customer_id=customer_id, doc_type=doc_type,
                    file_path=str(dest_path), file_name=doc_file.filename or dest_name,
                    mime_hint=mime_hint))
        else:
            db.add(CustomerDocument(
                customer_id=customer_id, doc_type="other", label=label.strip(),
                file_path=str(dest_path), file_name=doc_file.filename or dest_name,
                mime_hint=mime_hint))
        try:
            db.commit()
        except Exception:
            # Commit fail → xóa file mới vừa ghi để tránh orphan
            try:
                dest_path.unlink(missing_ok=True)
            except OSError:
                pass
            raise
    finally:
        db.close()

    # Commit thành công → xóa file cũ nếu có
    if old_file_to_delete and os.path.exists(old_file_to_delete):
        try:
            os.remove(old_file_to_delete)
        except OSError:
            pass  # Không chặn flow nếu xóa file cũ thất bại

    label_display = _DOC_LABELS.get(doc_type, doc_type)
    return _flash_redirect(f"/customers/{customer_id}",
                           f"Đã upload {label_display}", "success")


@router.get("/{customer_id}/docs/{doc_id}/view")
def view_doc(customer_id: int, doc_id: int, dl: int = 0):
    """Phục vụ file giấy tờ — inline (mặc định) hoặc download (?dl=1)."""
    db = SessionLocal()
    try:
        doc = _get_doc_or_404(db, customer_id, doc_id)
        path, name = doc.file_path, doc.file_name
    finally:
        db.close()

    if not os.path.exists(path):
        raise HTTPException(404, "File không còn tồn tại trên đĩa")

    _mt_map = {
        ".pdf":  "application/pdf",
        ".jpg":  "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png":  "image/png",
    }
    ext = os.path.splitext(path)[1].lower()
    media_type = _mt_map.get(ext, "application/octet-stream")

    return FileResponse(
        path=path, filename=name, media_type=media_type,
        content_disposition_type="attachment" if dl else "inline",
    )


@router.post("/{customer_id}/docs/{doc_id}/delete")
def delete_doc(customer_id: int, doc_id: int):
    """Xóa file giấy tờ — disk + DB record."""
    db = SessionLocal()
    try:
        doc = _get_doc_or_404(db, customer_id, doc_id)
        label = doc.label or _DOC_LABELS.get(doc.doc_type, doc.doc_type)
        if os.path.exists(doc.file_path):
            try:
                os.remove(doc.file_path)
            except OSError:
                pass
        db.delete(doc)
        db.commit()
    finally:
        db.close()
    return _flash_redirect(f"/customers/{customer_id}",
                           f"Đã xóa giấy tờ: {label}", "success")
